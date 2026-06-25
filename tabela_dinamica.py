import sys
import os
import sqlite3
import datetime
import re
import hashlib
import shutil

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget,
    QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox,
    QRadioButton, QButtonGroup, QGroupBox, QFileDialog,
    QMessageBox, QTableWidget, QTableWidgetItem,
    QTreeWidget, QTreeWidgetItem, QHeaderView, QSplitter,
    QAbstractItemView, QStatusBar, QFrame,
    QMenu, QWidgetAction, QDateEdit, QListWidget, QListWidgetItem,
    QInputDialog, QDialog, QDialogButtonBox,
)
from PyQt5.QtCore import Qt, QSize, QSortFilterProxyModel, QDate
from PyQt5.QtGui import QColor, QBrush, QFont

def _app_dir() -> str:
    """Retorna a pasta do .exe (quando compilado) ou do .py (em desenvolvimento)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

DB_PATH  = os.path.join(_app_dir(), "dados.db")
CFG_PATH = os.path.join(_app_dir(), "config.json")


def cfg_load() -> dict:
    try:
        import json
        with open(CFG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def cfg_save(data: dict):
    import json
    try:
        existing = cfg_load()
        existing.update(data)
        with open(CFG_PATH, "w", encoding="utf-8") as f:
            json.dump(existing, f, indent=2)
    except Exception:
        pass

NOMES_MESES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio",    6: "Junho",     7: "Julho", 8: "Agosto",
    9: "Setembro",10: "Outubro",  11: "Novembro", 12: "Dezembro",
}

COLOR_NEG   = QColor("#c62828")   # vermelho escuro
COLOR_POS   = QColor("#1b5e20")   # verde escuro
COLOR_ZERO  = QColor("#424242")   # cinza
BG_GRUPO    = QColor("#dce8f5")   # azul claro
BG_TOTAL    = QColor("#c8e6c9")   # verde claro
BG_ALT      = QColor("#f9f9f9")   # alternado

# ═══════════════════════════════════════════════════════════
#  AUTENTICAÇÃO (senha de acesso ao programa)
# ═══════════════════════════════════════════════════════════

def _hash_senha(senha: str, salt: str) -> str:
    return hashlib.sha256((salt + senha).encode("utf-8")).hexdigest()


class LoginDialog(QDialog):
    def __init__(self, erro=""):
        super().__init__()
        self.setWindowTitle("Acesso ao Programa")
        self.setModal(True)
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("Digite a senha para continuar:"))
        self._ed = QLineEdit()
        self._ed.setEchoMode(QLineEdit.Password)
        self._ed.returnPressed.connect(self.accept)
        lay.addWidget(self._ed)
        self._lbl_erro = QLabel(erro)
        self._lbl_erro.setStyleSheet("color:#c62828;")
        lay.addWidget(self._lbl_erro)
        botoes = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)
        lay.addWidget(botoes)
        self._ed.setFocus()

    def senha(self) -> str:
        return self._ed.text()


def _verificar_login() -> bool:
    """Retorna True se o acesso é permitido (sem senha configurada ou senha correta)."""
    cfg = cfg_load()
    auth = cfg.get("auth")
    if not auth:
        return True
    erro = ""
    while True:
        dlg = LoginDialog(erro)
        if dlg.exec_() != QDialog.Accepted:
            return False
        if _hash_senha(dlg.senha(), auth.get("salt", "")) == auth.get("hash", ""):
            return True
        erro = "Senha incorreta. Tente novamente."


# ═══════════════════════════════════════════════════════════
#  BANCO DE DADOS
# ═══════════════════════════════════════════════════════════

def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS registros (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            Data          TEXT,
            Mes           INTEGER,
            Ano           INTEGER,
            Categoria     TEXT,
            Sub_Categoria TEXT,
            Transacao     TEXT,
            Descricao     TEXT,
            Valor         REAL
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS categorias (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS subcategorias (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            nome         TEXT NOT NULL,
            categoria_id INTEGER NOT NULL REFERENCES categorias(id),
            UNIQUE(nome, categoria_id)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS transacoes (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        )
    """)
    con.commit()
    # garante que o banco nunca fica vazio (evita crash no pandas/Qt)
    cur = con.execute("SELECT COUNT(*) FROM registros")
    if cur.fetchone()[0] == 0:
        con.execute(
            "INSERT INTO registros (Data,Mes,Ano,Categoria,Sub_Categoria,Transacao,Descricao,Valor)"
            " VALUES (?,?,?,?,?,?,?,?)",
            ("01/01/1900", 1, 1900, "", "", "", "", 0.0)
        )
        con.commit()
    con.close()
    _migrar_cadastros_do_historico()


def _migrar_cadastros_do_historico():
    """Na 1ª vez que as tabelas de cadastro existirem vazias, recria os vínculos
    Categoria/Sub-Categoria/Transação a partir do que já foi usado nos lançamentos."""
    con = sqlite3.connect(DB_PATH)
    n_cat = con.execute("SELECT COUNT(*) FROM categorias").fetchone()[0]
    if n_cat > 0:
        con.close()
        return
    pares = con.execute(
        "SELECT DISTINCT Categoria, Sub_Categoria FROM registros"
        " WHERE Ano != 1900 AND Categoria != ''"
    ).fetchall()
    transacoes = con.execute(
        "SELECT DISTINCT Transacao FROM registros"
        " WHERE Ano != 1900 AND Transacao != ''"
    ).fetchall()
    cat_ids = {}
    for cat, sub in pares:
        if cat not in cat_ids:
            con.execute("INSERT OR IGNORE INTO categorias (nome) VALUES (?)", (cat,))
            cat_ids[cat] = con.execute(
                "SELECT id FROM categorias WHERE nome=?", (cat,)).fetchone()[0]
        if sub:
            con.execute(
                "INSERT OR IGNORE INTO subcategorias (nome, categoria_id) VALUES (?,?)",
                (sub, cat_ids[cat])
            )
    for (tr,) in transacoes:
        if tr:
            con.execute("INSERT OR IGNORE INTO transacoes (nome) VALUES (?)", (tr,))
    con.commit()
    con.close()


# ═══════════════════════════════════════════════════════════
#  CADASTROS: CATEGORIAS / SUB-CATEGORIAS / TRANSAÇÕES
# ═══════════════════════════════════════════════════════════

def listar_categorias():
    con = sqlite3.connect(DB_PATH)
    rows = con.execute("SELECT id, nome FROM categorias ORDER BY nome").fetchall()
    con.close()
    return rows


def id_categoria_por_nome(nome):
    con = sqlite3.connect(DB_PATH)
    row = con.execute("SELECT id FROM categorias WHERE nome=?", (nome,)).fetchone()
    con.close()
    return row[0] if row else None


def inserir_categoria(nome: str):
    con = sqlite3.connect(DB_PATH)
    try:
        con.execute("INSERT INTO categorias (nome) VALUES (?)", (nome.strip(),))
        con.commit()
        return True, ""
    except sqlite3.IntegrityError:
        return False, "Já existe uma categoria com esse nome."
    finally:
        con.close()


def renomear_categoria(cid: int, novo_nome: str):
    novo_nome = novo_nome.strip()
    con = sqlite3.connect(DB_PATH)
    old = con.execute("SELECT nome FROM categorias WHERE id=?", (cid,)).fetchone()[0]
    try:
        con.execute("UPDATE categorias SET nome=? WHERE id=?", (novo_nome, cid))
        con.execute("UPDATE registros SET Categoria=? WHERE Categoria=?", (novo_nome, old))
        con.commit()
        return True, ""
    except sqlite3.IntegrityError:
        con.rollback()
        return False, "Já existe uma categoria com esse nome."
    finally:
        con.close()


def excluir_categoria(cid: int):
    con = sqlite3.connect(DB_PATH)
    nome = con.execute("SELECT nome FROM categorias WHERE id=?", (cid,)).fetchone()[0]
    n_uso = con.execute(
        "SELECT COUNT(*) FROM registros WHERE Categoria=?", (nome,)).fetchone()[0]
    n_sub = con.execute(
        "SELECT COUNT(*) FROM subcategorias WHERE categoria_id=?", (cid,)).fetchone()[0]
    if n_uso:
        con.close()
        return False, f"Não é possível excluir: {n_uso} registro(s) usam esta categoria."
    if n_sub:
        con.close()
        return False, (f"Não é possível excluir: existem {n_sub} sub-categoria(s) "
                        f"cadastradas nesta categoria. Exclua-as primeiro.")
    con.execute("DELETE FROM categorias WHERE id=?", (cid,))
    con.commit()
    con.close()
    return True, ""


def listar_subcategorias(categoria_id: int):
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT id, nome FROM subcategorias WHERE categoria_id=? ORDER BY nome",
        (categoria_id,)).fetchall()
    con.close()
    return rows


def inserir_subcategoria(nome: str, categoria_id: int):
    con = sqlite3.connect(DB_PATH)
    try:
        con.execute(
            "INSERT INTO subcategorias (nome, categoria_id) VALUES (?,?)",
            (nome.strip(), categoria_id))
        con.commit()
        return True, ""
    except sqlite3.IntegrityError:
        return False, "Já existe uma sub-categoria com esse nome nesta categoria."
    finally:
        con.close()


def renomear_subcategoria(sid: int, novo_nome: str):
    novo_nome = novo_nome.strip()
    con = sqlite3.connect(DB_PATH)
    old, categoria_id = con.execute(
        "SELECT nome, categoria_id FROM subcategorias WHERE id=?", (sid,)).fetchone()
    cat_nome = con.execute(
        "SELECT nome FROM categorias WHERE id=?", (categoria_id,)).fetchone()[0]
    try:
        con.execute("UPDATE subcategorias SET nome=? WHERE id=?", (novo_nome, sid))
        con.execute(
            "UPDATE registros SET Sub_Categoria=? WHERE Sub_Categoria=? AND Categoria=?",
            (novo_nome, old, cat_nome))
        con.commit()
        return True, ""
    except sqlite3.IntegrityError:
        con.rollback()
        return False, "Já existe uma sub-categoria com esse nome nesta categoria."
    finally:
        con.close()


def excluir_subcategoria(sid: int):
    con = sqlite3.connect(DB_PATH)
    nome, categoria_id = con.execute(
        "SELECT nome, categoria_id FROM subcategorias WHERE id=?", (sid,)).fetchone()
    cat_nome = con.execute(
        "SELECT nome FROM categorias WHERE id=?", (categoria_id,)).fetchone()[0]
    n_uso = con.execute(
        "SELECT COUNT(*) FROM registros WHERE Sub_Categoria=? AND Categoria=?",
        (nome, cat_nome)).fetchone()[0]
    if n_uso:
        con.close()
        return False, f"Não é possível excluir: {n_uso} registro(s) usam esta sub-categoria."
    con.execute("DELETE FROM subcategorias WHERE id=?", (sid,))
    con.commit()
    con.close()
    return True, ""


def listar_transacoes():
    con = sqlite3.connect(DB_PATH)
    rows = con.execute("SELECT id, nome FROM transacoes ORDER BY nome").fetchall()
    con.close()
    return rows


def inserir_transacao(nome: str):
    con = sqlite3.connect(DB_PATH)
    try:
        con.execute("INSERT INTO transacoes (nome) VALUES (?)", (nome.strip(),))
        con.commit()
        return True, ""
    except sqlite3.IntegrityError:
        return False, "Já existe uma transação com esse nome."
    finally:
        con.close()


def renomear_transacao(tid: int, novo_nome: str):
    novo_nome = novo_nome.strip()
    con = sqlite3.connect(DB_PATH)
    old = con.execute("SELECT nome FROM transacoes WHERE id=?", (tid,)).fetchone()[0]
    try:
        con.execute("UPDATE transacoes SET nome=? WHERE id=?", (novo_nome, tid))
        con.execute("UPDATE registros SET Transacao=? WHERE Transacao=?", (novo_nome, old))
        con.commit()
        return True, ""
    except sqlite3.IntegrityError:
        con.rollback()
        return False, "Já existe uma transação com esse nome."
    finally:
        con.close()


def excluir_transacao(tid: int):
    con = sqlite3.connect(DB_PATH)
    nome = con.execute("SELECT nome FROM transacoes WHERE id=?", (tid,)).fetchone()[0]
    n_uso = con.execute(
        "SELECT COUNT(*) FROM registros WHERE Transacao=?", (nome,)).fetchone()[0]
    if n_uso:
        con.close()
        return False, f"Não é possível excluir: {n_uso} registro(s) usam esta transação."
    con.execute("DELETE FROM transacoes WHERE id=?", (tid,))
    con.commit()
    con.close()
    return True, ""


def _mes_ano(data_str: str):
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            d = datetime.datetime.strptime(data_str.strip(), fmt)
            return d.month, d.year
        except ValueError:
            pass
    return None, None


def _parse_data(data_str: str):
    """Converte a string de data para datetime; None se não reconhecer."""
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(str(data_str).strip(), fmt)
        except ValueError:
            pass
    return None


def _parse_valor(raw) -> float:
    """Converte texto em número, aceitando tanto ',' quanto '.' como separador decimal."""
    s = re.sub(r"[R$\s]", "", str(raw).strip())
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rindex(",") > s.rindex("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    return float(s)


def inserir(row: dict):
    mes, ano = _mes_ano(row.get("Data", ""))
    con = sqlite3.connect(DB_PATH)
    cur = con.execute(
        "INSERT INTO registros (Data,Mes,Ano,Categoria,Sub_Categoria,Transacao,Descricao,Valor)"
        " VALUES (?,?,?,?,?,?,?,?)",
        (row["Data"], mes, ano, row["Categoria"], row["Sub_Categoria"],
         row["Transacao"], row["Descricao"], _parse_valor(row["Valor"]) if row["Valor"] else 0.0)
    )
    new_id = cur.lastrowid
    con.commit()
    con.close()
    return new_id


def atualizar(rid, row: dict):
    mes, ano = _mes_ano(row.get("Data", ""))
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "UPDATE registros SET Data=?,Mes=?,Ano=?,Categoria=?,Sub_Categoria=?,"
        "Transacao=?,Descricao=?,Valor=? WHERE id=?",
        (row["Data"], mes, ano, row["Categoria"], row["Sub_Categoria"],
         row["Transacao"], row["Descricao"], _parse_valor(row["Valor"]) if row["Valor"] else 0.0, rid)
    )
    con.commit()
    con.close()


def _inserir_dummy(con):
    con.execute(
        "INSERT INTO registros (Data,Mes,Ano,Categoria,Sub_Categoria,Transacao,Descricao,Valor)"
        " VALUES (?,?,?,?,?,?,?,?)",
        ("01/01/1900", 1, 1900, "", "", "", "", 0.0)
    )


def deletar(rid):
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM registros WHERE id=?", (rid,))
    con.commit()
    cur = con.execute("SELECT COUNT(*) FROM registros")
    if cur.fetchone()[0] == 0:
        _inserir_dummy(con)
        con.commit()
    con.close()


def buscar_todos():
    con = sqlite3.connect(DB_PATH)
    cur = con.execute("SELECT * FROM registros ORDER BY Data")
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description]
    con.close()
    return cols, rows


def buscar_ultimo_registro():
    """Retorna o registro real (não dummy) mais recentemente inserido, ou None."""
    con = sqlite3.connect(DB_PATH)
    cur = con.execute(
        "SELECT Data,Categoria,Sub_Categoria,Transacao FROM registros"
        " WHERE Ano != 1900 ORDER BY id DESC LIMIT 1")
    row = cur.fetchone()
    con.close()
    if not row:
        return None
    return {"Data": row[0], "Categoria": row[1], "Sub_Categoria": row[2], "Transacao": row[3]}


def apagar_banco():
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM registros")
    _inserir_dummy(con)
    con.commit()
    con.close()


class _ResolvedorCadastros:
    """Resolve Categoria/Sub-Categoria/Transação durante a importação:
    se já existir um cadastro com a mesma grafia (ignorando maiúsculas/minúsculas
    e acentuação via casefold), reaproveita a grafia já cadastrada; caso contrário,
    cria um novo cadastro com o texto exatamente como veio na importação."""

    def __init__(self, con):
        self.con = con
        self._cat_cf = {}      # casefold(nome) -> (id, nome canônico)
        self._sub_cf = {}      # (categoria_id, casefold(nome)) -> nome canônico
        self._tran_cf = {}     # casefold(nome) -> nome canônico
        for cid, nome in con.execute("SELECT id, nome FROM categorias"):
            self._cat_cf[nome.casefold()] = (cid, nome)
        for nome, cid in con.execute("SELECT nome, categoria_id FROM subcategorias"):
            self._sub_cf[(cid, nome.casefold())] = nome
        for (nome,) in con.execute("SELECT nome FROM transacoes"):
            self._tran_cf[nome.casefold()] = nome

    def categoria(self, nome: str):
        nome = nome.strip()
        if not nome:
            return "", None
        key = nome.casefold()
        if key in self._cat_cf:
            cid, canonico = self._cat_cf[key]
            return canonico, cid
        cur = self.con.execute("INSERT INTO categorias (nome) VALUES (?)", (nome,))
        cid = cur.lastrowid
        self._cat_cf[key] = (cid, nome)
        return nome, cid

    def subcategoria(self, nome: str, categoria_id):
        nome = nome.strip()
        if not nome or categoria_id is None:
            return nome
        key = (categoria_id, nome.casefold())
        if key in self._sub_cf:
            return self._sub_cf[key]
        self.con.execute(
            "INSERT INTO subcategorias (nome, categoria_id) VALUES (?,?)",
            (nome, categoria_id))
        self._sub_cf[key] = nome
        return nome

    def transacao(self, nome: str):
        nome = nome.strip()
        if not nome:
            return ""
        key = nome.casefold()
        if key in self._tran_cf:
            return self._tran_cf[key]
        self.con.execute("INSERT INTO transacoes (nome) VALUES (?)", (nome,))
        self._tran_cf[key] = nome
        return nome


def importar_df(df, modo: str):
    col_map = {c.lower().strip(): c for c in df.columns}

    def get(*names):
        for name in names:
            key = name.lower().strip()
            if key in col_map:
                return col_map[key]
        prefix = names[0].lower().strip()
        for k, v in col_map.items():
            if k.startswith(prefix):
                return v
        return None

    if modo == "sobrescrever":
        apagar_banco()

    col_data  = get("data e hora", "data")
    col_cat   = get("categoria")
    col_sub   = get("sub_categoria", "sub-categoria", "subcategoria")
    col_tran  = get("transação", "transacao", "transação")
    col_desc  = get("descrição", "descricao", "descrição")
    col_valor = get("valor")

    con = sqlite3.connect(DB_PATH)
    resolvedor = _ResolvedorCadastros(con)
    for _, r in df.iterrows():
        data_val = str(r[col_data]).strip() if col_data else ""
        mes, ano = _mes_ano(data_val)
        if col_valor:
            try:
                valor = _parse_valor(r[col_valor])
            except ValueError:
                valor = 0.0
        else:
            valor = 0.0

        cat_raw  = str(r[col_cat]).strip()  if col_cat  else ""
        sub_raw  = str(r[col_sub]).strip()  if col_sub  else ""
        tran_raw = str(r[col_tran]).strip() if col_tran else ""

        cat_final, cat_id = resolvedor.categoria(cat_raw)
        sub_final = resolvedor.subcategoria(sub_raw, cat_id)
        tran_final = resolvedor.transacao(tran_raw)

        con.execute(
            "INSERT INTO registros (Data,Mes,Ano,Categoria,Sub_Categoria,Transacao,Descricao,Valor)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (data_val, mes, ano,
             cat_final,
             sub_final,
             tran_final,
             str(r[col_desc]) if col_desc else "",
             valor)
        )
    con.commit()
    con.close()


# ═══════════════════════════════════════════════════════════
#  HELPERS VISUAIS
# ═══════════════════════════════════════════════════════════

def fmt_valor(v) -> str:
    if not isinstance(v, (int, float)):
        return str(v)
    s = f"R$ {abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"-{s}" if v < 0 else s


def cor_valor(v) -> QColor:
    try:
        f = float(v) if not isinstance(v, float) else v
        if f < 0:   return COLOR_NEG
        if f > 0:   return COLOR_POS
        return COLOR_ZERO
    except (ValueError, TypeError):
        return COLOR_ZERO


class NumericItem(QTableWidgetItem):
    """Item que ordena pelo valor numérico armazenado em UserRole."""
    def __lt__(self, other):
        a = self.data(Qt.UserRole)
        b = other.data(Qt.UserRole)
        try:
            return float(a) < float(b)
        except (TypeError, ValueError):
            return super().__lt__(other)


def item_valor(v) -> QTableWidgetItem:
    it = NumericItem(fmt_valor(v))
    it.setData(Qt.UserRole, float(v) if v is not None else 0.0)
    it.setForeground(QBrush(cor_valor(v)))
    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
    return it


def _btn(texto, cor_bg, slot, min_w=100):
    """Cria um QPushButton padronizado."""
    b = QPushButton(texto)
    b.setStyleSheet(
        f"QPushButton{{background:{cor_bg};color:white;border-radius:5px;"
        f"padding:6px 18px;font-weight:bold;font-size:12px;}}"
        f"QPushButton:hover{{background:{cor_bg};border:1px solid rgba(0,0,0,0.2);}}"
    )
    b.setMinimumWidth(min_w)
    b.clicked.connect(slot)
    return b


# ═══════════════════════════════════════════════════════════
#  ABA DADOS
# ═══════════════════════════════════════════════════════════

COLS_DADOS = ["id", "Data", "Mês", "Ano", "Categoria", "Sub-Categoria",
              "Transação", "Descrição", "Valor"]

class AbaForm(QWidget):
    def __init__(self):
        super().__init__()
        self._edit_id  = None
        self._all_rows = []   # cache completo para filtro local
        self._dirty    = set()  # campos modificados pelo usuário desde última seleção
        self._auto_ajuste_feito = False  # auto-ajuste de colunas só na 1ª carga
        self._build()

    # ── construção ────────────────────────────────────────
    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(10, 8, 10, 6)
        root.setSpacing(6)

        # ── formulário ────────────────────────────────────
        grp = QGroupBox("Registro")
        form = QGridLayout(grp)
        form.setSpacing(5)
        form.setContentsMargins(8, 6, 8, 6)

        self._campos = {}
        specs = [
            ("Data",          "Data (dd/mm/aaaa hh:mm:ss)", False),
            ("Categoria",     "Categoria",                   True),
            ("Sub_Categoria", "Sub-Categoria",               True),
            ("Transacao",     "Transação",                   True),
            ("Descricao",     "Descrição",                   False),
            ("Valor",         "Valor",                       False),
        ]
        for row_idx, (key, lbl_txt, is_combo) in enumerate(specs):
            form.addWidget(QLabel(lbl_txt + ":"), row_idx, 0, Qt.AlignRight)
            if is_combo:
                # combos de Categoria/Sub-Categoria/Transação só permitem
                # escolher valores já cadastrados (aba "Categorias") —
                # evita categorias "sujas" criadas por erro de digitação
                w = QComboBox()
                w.setEditable(False)
                w.setMinimumWidth(260)
            else:
                w = QLineEdit()
                if key == "Descricao":
                    w.setMinimumWidth(450)
            form.addWidget(w, row_idx, 1, Qt.AlignLeft)
            self._campos[key] = w
        # ao trocar a Categoria, a lista de Sub-Categoria deve refletir só
        # as sub-categorias cadastradas para aquela categoria
        self._campos["Categoria"].currentTextChanged.connect(self._on_categoria_form_changed)

        # botões do formulário
        btn_row = QHBoxLayout()
        self._btn_salvar  = _btn("Salvar",             "#4CAF50", self._salvar,  110)
        self._btn_limpar  = _btn("Limpar",             "#2196F3", self._limpar,  110)
        self._btn_excluir = _btn("Excluir Selecionado","#f44336", self._excluir, 150)
        self._btn_dup     = _btn("Duplicar Selecionado","#00897B", self._duplicar, 150)
        self._btn_lote    = _btn("Aplicar a Selecionados", "#E65100", self._aplicar_lote, 160)
        self._btn_lote.setVisible(False)
        self._btn_dup.setVisible(False)
        btn_row.addWidget(self._btn_salvar)
        btn_row.addWidget(self._btn_limpar)
        btn_row.addWidget(self._btn_excluir)
        btn_row.addWidget(self._btn_dup)
        btn_row.addWidget(self._btn_lote)
        btn_row.addStretch()
        form.addLayout(btn_row, len(specs), 0, 1, 2)
        root.addWidget(grp)

        # conectar dirty tracking em todos os campos
        for key, w in self._campos.items():
            if isinstance(w, QComboBox):
                w.currentTextChanged.connect(lambda _, k=key: self._marcar_dirty(k))
            else:
                w.textChanged.connect(lambda _, k=key: self._marcar_dirty(k))

        # ── barra de filtros ──────────────────────────────
        flt_grp = QGroupBox("Filtros / Pesquisa")
        flt_lay = QHBoxLayout(flt_grp)
        flt_lay.setSpacing(6)
        flt_lay.setContentsMargins(8, 4, 8, 4)

        self._flt_widgets = {}
        flt_specs = [
            ("data", "Data",          110),
            ("cat",  "Categoria",     120),
            ("sub",  "Sub-Categoria", 120),
            ("tran", "Transação",     150),
            ("desc", "Descrição",     160),
        ]
        for key, placeholder, w_px in flt_specs:
            ed = QLineEdit()
            ed.setPlaceholderText(placeholder)
            ed.setFixedWidth(w_px)
            ed.textChanged.connect(self._aplicar_filtro)
            flt_lay.addWidget(ed)
            self._flt_widgets[key] = ed

        flt_lay.addWidget(QLabel("Ano:"))
        self._flt_ano = QComboBox(); self._flt_ano.setFixedWidth(90)
        self._flt_ano.currentTextChanged.connect(self._aplicar_filtro)
        flt_lay.addWidget(self._flt_ano)

        flt_lay.addWidget(QLabel("Mês:"))
        self._flt_mes = QComboBox(); self._flt_mes.setFixedWidth(130)
        self._flt_mes.currentTextChanged.connect(self._aplicar_filtro)
        flt_lay.addWidget(self._flt_mes)

        btn_limpar_flt = _btn("✕ Limpar", "#757575", self._limpar_filtros, 80)
        flt_lay.addWidget(btn_limpar_flt)
        flt_lay.addStretch()
        root.addWidget(flt_grp)

        # ── tabela ────────────────────────────────────────
        self._table = QTableWidget(0, len(COLS_DADOS))
        self._table.setHorizontalHeaderLabels(COLS_DADOS)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        # manter destaque da seleção mesmo quando a tabela não tem foco
        self._table.setStyleSheet("""
            QTableWidget::item:selected {
                background-color: #1976D2;
                color: white;
            }
            QTableWidget::item:selected:!active {
                background-color: #1976D2;
                color: white;
            }
        """)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        hdr.setStretchLastSection(False)
        hdr.setSortIndicatorShown(True)
        self._table.setSortingEnabled(True)
        self._table.itemSelectionChanged.connect(self._on_select)
        # salva larguras automaticamente quando o usuário arrastar uma coluna
        hdr.sectionResized.connect(self._salvar_layout)
        root.addWidget(self._table, 1)

        # ── rodapé: status ────────────────────────────────
        self._status = QLabel("")
        self._status.setStyleSheet("color:#555; font-size:10px")
        root.addWidget(self._status)

        # ── exportação da aba Dados ───────────────────────
        exp_row = QHBoxLayout()
        exp_row.addStretch()
        exp_row.addWidget(_btn("Exportar XLSX", "#1565C0", self._exportar_xlsx, 130))
        exp_row.addWidget(_btn("Exportar CSV",  "#6A1B9A", self._exportar_csv,  120))
        root.addLayout(exp_row)

        self._carregar()

    # ── campos helpers ────────────────────────────────────
    def _get_text(self, key) -> str:
        w = self._campos[key]
        if isinstance(w, QComboBox):
            return w.currentText().strip()
        return w.text().strip()

    def _set_text(self, key, val):
        w = self._campos[key]
        if isinstance(w, QComboBox):
            idx = w.findText(val)
            if idx >= 0:
                w.setCurrentIndex(idx)
            else:
                w.setCurrentText(val)
        else:
            w.setText(val)

    def _clear_field(self, key):
        w = self._campos[key]
        if isinstance(w, QComboBox):
            w.setCurrentIndex(-1)
            w.clearEditText()
        else:
            w.clear()

    def _atualizar_combos(self):
        """Recarrega Categoria/Sub-Categoria/Transação a partir dos cadastros
        (aba "Categorias"), mantendo a seleção atual quando ainda existir."""
        cb_cat = self._campos["Categoria"]
        cur_cat = cb_cat.currentText()
        cats = [n for _, n in listar_categorias()]
        cb_cat.blockSignals(True)
        cb_cat.clear()
        cb_cat.addItems(cats)
        idx = cb_cat.findText(cur_cat)
        cb_cat.setCurrentIndex(idx if idx >= 0 else (0 if cats else -1))
        cb_cat.blockSignals(False)

        self._atualizar_subcats_form(preservar=True)

        cb_tran = self._campos["Transacao"]
        cur_tran = cb_tran.currentText()
        trans = [n for _, n in listar_transacoes()]
        cb_tran.blockSignals(True)
        cb_tran.clear()
        cb_tran.addItems(trans)
        idx = cb_tran.findText(cur_tran)
        cb_tran.setCurrentIndex(idx if idx >= 0 else (0 if trans else -1))
        cb_tran.blockSignals(False)

    def _atualizar_subcats_form(self, preservar=True):
        """Recarrega o combo de Sub-Categoria com apenas as sub-categorias
        cadastradas para a Categoria atualmente selecionada no formulário."""
        cb_cat = self._campos["Categoria"]
        cb_sub = self._campos["Sub_Categoria"]
        cur_sub = cb_sub.currentText() if preservar else ""
        cat_nome = cb_cat.currentText()
        subs = []
        if cat_nome:
            cid = id_categoria_por_nome(cat_nome)
            if cid:
                subs = [n for _, n in listar_subcategorias(cid)]
        cb_sub.blockSignals(True)
        cb_sub.clear()
        cb_sub.addItems(subs)
        idx = cb_sub.findText(cur_sub)
        cb_sub.setCurrentIndex(idx if idx >= 0 else (0 if subs else -1))
        cb_sub.blockSignals(False)

    def _on_categoria_form_changed(self, _texto):
        if getattr(self, "_carregando_selecao", False):
            return
        self._atualizar_subcats_form(preservar=False)

    def _atualizar_filtros_combo(self):
        anos  = ["(todos)"] + sorted(set(
            str(r[3]) for r in self._all_rows if r[3] is not None))
        meses = ["(todos)"] + [f"{i} – {NOMES_MESES[i]}" for i in range(1, 13)]
        for cb, vals in ((self._flt_ano, anos), (self._flt_mes, meses)):
            cur = cb.currentText()
            cb.blockSignals(True)
            cb.clear(); cb.addItems(vals)
            idx = cb.findText(cur)
            cb.setCurrentIndex(idx if idx >= 0 else 0)
            cb.blockSignals(False)

    # ── ações formulário ──────────────────────────────────
    def _salvar(self):
        row = {k: self._get_text(k) for k in self._campos}
        if not row["Data"]:
            QMessageBox.warning(self, "Atenção", "Data é obrigatória.")
            return
        mes, ano = _mes_ano(row["Data"])
        if mes is None:
            QMessageBox.warning(self, "Atenção",
                "Formato de data inválido.\nUse dd/mm/aaaa ou dd/mm/aaaa hh:mm:ss")
            return
        try:
            valor = _parse_valor(row["Valor"]) if row["Valor"] else 0.0
        except ValueError:
            QMessageBox.warning(self, "Atenção", "Valor deve ser numérico.")
            return
        row["Valor"] = valor
        edit_id = self._edit_id
        if edit_id:
            atualizar(edit_id, row)
            nova = (edit_id, row["Data"], mes, ano, row["Categoria"],
                    row["Sub_Categoria"], row["Transacao"], row["Descricao"], valor)
            self._all_rows = [nova if r[0] == edit_id else r
                              for r in self._all_rows]
        else:
            new_id = inserir(row)
            nova = (new_id, row["Data"], mes, ano, row["Categoria"],
                    row["Sub_Categoria"], row["Transacao"], row["Descricao"], valor)
            self._all_rows.append(nova)
        # combos do formulário e dos filtros (cálculo barato)
        self._atualizar_combos()
        self._atualizar_filtros_combo()
        # atualização incremental da tabela (só a linha afetada)
        if edit_id:
            self._editar_linha_tabela(edit_id, nova)
        else:
            self._inserir_linha_tabela(nova)
        self._limpar()

    def _linha_por_id(self, rid):
        """Índice da linha na tabela cujo id (coluna 0) == rid, ou -1."""
        for i in range(self._table.rowCount()):
            it = self._table.item(i, 0)
            if it and it.text() == str(rid):
                return i
        return -1

    def _reordenar_tabela(self):
        hdr = self._table.horizontalHeader()
        self._table.sortItems(hdr.sortIndicatorSection(), hdr.sortIndicatorOrder())

    def _inserir_linha_tabela(self, r):
        """Adiciona uma única linha nova à tabela (se passar no filtro)."""
        if self._passa_filtro(r):
            self._table.setSortingEnabled(False)
            i = self._table.rowCount()
            self._table.insertRow(i)
            self._montar_item_linha(i, r)
            self._table.setSortingEnabled(True)
            self._reordenar_tabela()
        self._atualizar_soma_status()

    def _editar_linha_tabela(self, rid, r):
        """Atualiza in-place a linha editada (ou insere/remove se mudou o filtro)."""
        i = self._linha_por_id(rid)
        passa = self._passa_filtro(r)
        self._table.setSortingEnabled(False)
        if i >= 0 and passa:
            self._montar_item_linha(i, r)        # atualiza só essa linha
        elif i >= 0 and not passa:
            self._table.removeRow(i)             # saiu do filtro
        elif i < 0 and passa:
            j = self._table.rowCount()
            self._table.insertRow(j)
            self._montar_item_linha(j, r)        # entrou no filtro
        self._table.setSortingEnabled(True)
        if passa:
            self._reordenar_tabela()
        self._atualizar_soma_status()

    def _limpar(self):
        self._carregando_selecao = True
        for key in self._campos:
            self._clear_field(key)
        ultimo = buscar_ultimo_registro()
        if ultimo:
            self._set_text("Data", ultimo["Data"] or "")
            self._set_text("Categoria", ultimo["Categoria"] or "")
            self._atualizar_subcats_form(preservar=False)
            self._set_text("Sub_Categoria", ultimo["Sub_Categoria"] or "")
            self._set_text("Transacao", ultimo["Transacao"] or "")
        self._carregando_selecao = False
        self._edit_id = None
        self._dirty.clear()
        self._btn_salvar.setText("Salvar")
        self._btn_salvar.setVisible(True)
        self._btn_lote.setVisible(False)

    def _excluir(self):
        sel = self._table.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "Info", "Selecione um ou mais registros para excluir.")
            return
        count = len(sel)
        msg = (f"Excluir {count} registros selecionados?"
               if count > 1 else "Excluir registro selecionado?")
        if QMessageBox.question(self, "Confirmar", msg,
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            rids = [int(self._table.item(idx.row(), 0).text()) for idx in sel]
            linhas = sorted((idx.row() for idx in sel), reverse=True)
            for rid in rids:
                deletar(rid)
            rids_set = set(rids)
            self._all_rows = [r for r in self._all_rows if r[0] not in rids_set]
            # se o banco ficou vazio, deletar() reinsere o registro dummy:
            # nesse caso faz uma recarga completa para refleti-lo
            if not self._all_rows:
                self._limpar()
                self._carregar()
                return
            # remoção incremental das linhas selecionadas (de baixo p/ cima)
            self._table.setSortingEnabled(False)
            for r in linhas:
                self._table.removeRow(r)
            self._table.setSortingEnabled(True)
            self._atualizar_combos()
            self._atualizar_filtros_combo()
            self._atualizar_soma_status()
            self._limpar()

    def _marcar_dirty(self, key):
        # ignora mudanças causadas programaticamente durante _on_select
        if getattr(self, "_carregando_selecao", False):
            return
        self._dirty.add(key)

    def _on_select(self):
        sel = self._table.selectionModel().selectedRows()
        if not sel:
            self._btn_lote.setVisible(False)
            self._btn_dup.setVisible(False)
            return
        multi = len(sel) > 1
        self._btn_lote.setVisible(multi)
        self._btn_dup.setVisible(not multi)
        self._btn_salvar.setVisible(not multi)

        # carrega dados da primeira linha selecionada no formulário
        self._carregando_selecao = True
        self._dirty.clear()
        r = sel[0].row()
        self._edit_id = int(self._table.item(r, 0).text())
        self._set_text("Data",          self._table.item(r, 1).text())
        self._set_text("Categoria",     self._table.item(r, 4).text())
        self._atualizar_subcats_form(preservar=False)
        self._set_text("Sub_Categoria", self._table.item(r, 5).text())
        self._set_text("Transacao",     self._table.item(r, 6).text())
        self._set_text("Descricao",     self._table.item(r, 7).text())
        raw = self._table.item(r, 8).data(Qt.UserRole)
        self._set_text("Valor", str(raw) if raw is not None else "")
        self._carregando_selecao = False
        self._btn_salvar.setText("Atualizar" if not multi else "Salvar")

    def _duplicar(self):
        sel = self._table.selectionModel().selectedRows()
        if not sel:
            return
        r = sel[0].row()
        self._carregando_selecao = True
        self._set_text("Data",          self._table.item(r, 1).text())
        self._set_text("Categoria",     self._table.item(r, 4).text())
        self._atualizar_subcats_form(preservar=False)
        self._set_text("Sub_Categoria", self._table.item(r, 5).text())
        self._set_text("Transacao",     self._table.item(r, 6).text())
        self._set_text("Descricao",     self._table.item(r, 7).text())
        raw = self._table.item(r, 8).data(Qt.UserRole)
        self._set_text("Valor", str(raw) if raw is not None else "")
        self._carregando_selecao = False
        self._edit_id = None
        self._dirty.clear()
        self._btn_salvar.setText("Salvar")
        self._btn_salvar.setVisible(True)
        self._table.clearSelection()

    def _aplicar_lote(self):
        sel = self._table.selectionModel().selectedRows()
        if not sel or not self._dirty:
            QMessageBox.information(self, "Info",
                "Altere pelo menos um campo antes de aplicar.")
            return
        rids   = [int(self._table.item(idx.row(), 0).text()) for idx in sel]
        campos = sorted(self._dirty)
        nomes  = {"Data": "Data", "Categoria": "Categoria",
                  "Sub_Categoria": "Sub-Categoria", "Transacao": "Transação",
                  "Descricao": "Descrição", "Valor": "Valor"}
        lista  = ", ".join(nomes.get(c, c) for c in campos)
        msg    = (f"Atualizar o(s) campo(s)  【{lista}】\n"
                  f"em {len(rids)} registros selecionados?")
        if QMessageBox.question(self, "Confirmar edição em lote", msg,
                                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        # busca registro completo e substitui apenas os campos dirty
        con = sqlite3.connect(DB_PATH)
        for rid in rids:
            cur = con.execute(
                "SELECT Data,Categoria,Sub_Categoria,Transacao,Descricao,Valor"
                " FROM registros WHERE id=?", (rid,))
            row_db = cur.fetchone()
            if not row_db:
                continue
            row = dict(zip(
                ["Data","Categoria","Sub_Categoria","Transacao","Descricao","Valor"],
                row_db))
            for campo in campos:
                row[campo] = self._get_text(campo)
            mes, ano = _mes_ano(row["Data"])
            try:
                valor = _parse_valor(row["Valor"]) if row["Valor"] else 0.0
            except ValueError:
                valor = 0.0
            con.execute(
                "UPDATE registros SET Data=?,Mes=?,Ano=?,Categoria=?,Sub_Categoria=?,"
                "Transacao=?,Descricao=?,Valor=? WHERE id=?",
                (row["Data"], mes, ano, row["Categoria"], row["Sub_Categoria"],
                 row["Transacao"], row["Descricao"], valor, rid))
        con.commit()
        con.close()
        self._dirty.clear()
        self._carregar()

    # ── filtros ───────────────────────────────────────────
    def _limpar_filtros(self):
        for ed in self._flt_widgets.values():
            ed.blockSignals(True); ed.clear(); ed.blockSignals(False)
        self._flt_ano.blockSignals(True)
        self._flt_ano.setCurrentIndex(0)
        self._flt_ano.blockSignals(False)
        self._flt_mes.blockSignals(True)
        self._flt_mes.setCurrentIndex(0)
        self._flt_mes.blockSignals(False)
        self._aplicar_filtro()

    def _criterios_filtro(self):
        """Retorna os critérios de filtro atuais como tupla, para reuso."""
        f_data = self._flt_widgets["data"].text().lower()
        f_cat  = self._flt_widgets["cat"].text().lower()
        f_sub  = self._flt_widgets["sub"].text().lower()
        f_tran = self._flt_widgets["tran"].text().lower()
        f_desc = self._flt_widgets["desc"].text().lower()
        f_ano  = self._flt_ano.currentText()
        f_mes  = self._flt_mes.currentText()
        num_mes = int(f_mes.split(" – ")[0]) if f_mes != "(todos)" and " – " in f_mes else None
        num_ano = int(f_ano) if f_ano not in ("", "(todos)") else None
        return (f_data, f_cat, f_sub, f_tran, f_desc, num_ano, num_mes)

    def _passa_filtro(self, r, crit=None):
        """True se a linha r passa nos filtros atuais."""
        f_data, f_cat, f_sub, f_tran, f_desc, num_ano, num_mes = \
            crit if crit is not None else self._criterios_filtro()
        if f_data and f_data not in str(r[1]).lower(): return False
        if f_cat  and f_cat  not in str(r[4]).lower(): return False
        if f_sub  and f_sub  not in str(r[5]).lower(): return False
        if f_tran and f_tran not in str(r[6]).lower(): return False
        if f_desc and f_desc not in str(r[7]).lower(): return False
        if num_ano is not None and r[3] != num_ano:    return False
        if num_mes is not None and r[2] != num_mes:    return False
        return True

    def _montar_item_linha(self, i, r):
        """Cria e posiciona os 9 itens da linha i da tabela a partir de r."""
        for j, v in enumerate(r):
            if j == 8:                          # Valor
                it = item_valor(v)
            elif j in (0, 2, 3):               # id, Mês, Ano
                it = NumericItem(str(v) if v is not None else "")
                it.setData(Qt.UserRole, float(v) if v is not None else 0.0)
                it.setTextAlignment(Qt.AlignCenter)
            elif j == 1:                        # Data — ordena por ISO
                it = NumericItem(str(v) if v is not None else "")
                d = _parse_data(str(v)) if v else None
                sort_key = float(d.strftime("%Y%m%d%H%M%S")) if d else 0.0
                it.setData(Qt.UserRole, sort_key)
                it.setTextAlignment(Qt.AlignCenter)
            else:
                it = QTableWidgetItem(str(v) if v is not None else "")
                it.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(i, j, it)

    def _atualizar_soma_status(self):
        """Recalcula a soma (cabeçalho Valor) e o status, sobre as linhas visíveis."""
        crit = self._criterios_filtro()
        soma = 0.0
        vis = 0
        for r in self._all_rows:
            if self._passa_filtro(r, crit):
                soma += float(r[8] or 0)
                vis += 1
        cor_hex = "#c62828" if soma < 0 else "#1b5e20"
        hdr_item = self._table.horizontalHeaderItem(8)
        hdr_item.setText(f"Valor  |  {fmt_valor(soma)}")
        hdr_item.setForeground(QBrush(QColor(cor_hex)))
        total = len(self._all_rows)
        self._status.setText(
            f"Exibindo {vis} de {total} registros" +
            (f"  |  filtro ativo" if vis < total else ""))

    def _ajustar_larguras(self):
        """Restaura larguras salvas ou auto-ajusta apenas na 1ª carga da sessão."""
        hdr = self._table.horizontalHeader()
        hdr.sectionResized.disconnect(self._salvar_layout)
        saved = cfg_load().get("dados_col_widths")
        if saved and len(saved) == self._table.columnCount():
            # layout salvo pelo usuário: restaura direto, ignora auto-ajuste
            for c, w in enumerate(saved):
                self._table.setColumnWidth(c, w)
        elif not self._auto_ajuste_feito:
            # sem layout salvo: auto-ajuste por conteúdo + cabeçalho (uma vez)
            self._table.resizeColumnsToContents()
            fm = self._table.fontMetrics()
            for col in range(self._table.columnCount()):
                hdr_txt = self._table.horizontalHeaderItem(col)
                hdr_w = fm.horizontalAdvance(hdr_txt.text() if hdr_txt else "") + 24
                if hdr_w > hdr.sectionSize(col):
                    hdr.resizeSection(col, hdr_w)
            self._auto_ajuste_feito = True
        hdr.sectionResized.connect(self._salvar_layout)

    def _aplicar_filtro(self):
        crit = self._criterios_filtro()
        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)
        for r in self._all_rows:
            # r = (id, Data, Mes, Ano, Categoria, Sub_Categoria, Transacao, Descricao, Valor)
            if not self._passa_filtro(r, crit):
                continue
            i = self._table.rowCount()
            self._table.insertRow(i)
            self._montar_item_linha(i, r)
        self._table.setSortingEnabled(True)
        self._atualizar_soma_status()
        self._ajustar_larguras()

    # ── carga ─────────────────────────────────────────────
    def _carregar(self):
        _, rows = buscar_todos()
        self._all_rows = list(rows)
        self._auto_ajuste_feito = False   # recarga completa reavalia larguras
        self._atualizar_combos()
        self._atualizar_filtros_combo()
        self._aplicar_filtro()

    # ── layout de colunas ─────────────────────────────────
    def _salvar_layout(self):
        """Chamado automaticamente ao arrastar qualquer coluna."""
        widths = [self._table.columnWidth(c)
                  for c in range(self._table.columnCount())]
        cfg_save({"dados_col_widths": widths})


    # ── exportação ────────────────────────────────────────
    def _linhas_visiveis(self, tipado=False):
        """Retorna linhas visíveis; coluna Valor como número puro (sem R$).

        Se tipado=True, converte id/Mês/Ano para inteiro e Data para
        datetime, para que o Excel reconheça os campos como numéricos/data
        em vez de texto.
        """
        rows = []
        for i in range(self._table.rowCount()):
            row = []
            for j in range(self._table.columnCount()):
                it = self._table.item(i, j)
                if j == 8 and it:                          # coluna Valor
                    raw = it.data(Qt.UserRole)
                    row.append(raw if raw is not None else "")
                elif tipado and j in (0, 2, 3) and it:     # id, Mês, Ano → int
                    txt = it.text().strip()
                    try:
                        row.append(int(float(txt)) if txt != "" else "")
                    except ValueError:
                        row.append(txt)
                elif tipado and j == 1 and it:             # Data → datetime
                    row.append(_parse_data(it.text()) or it.text())
                else:
                    row.append(it.text() if it else "")
            rows.append(row)
        return rows

    def _exportar_xlsx(self):
        if not OPENPYXL_OK:
            QMessageBox.critical(self, "Erro",
                "openpyxl não instalado.\nExecute: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar como XLSX", "", "Excel (*.xlsx)")
        if not path:
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"
        # cabeçalho (sem a soma no texto)
        hdrs = [COLS_DADOS[i] if i != 8 else "Valor"
                for i in range(len(COLS_DADOS))]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(self._linhas_visiveis(tipado=True), 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(
                    horizontal="right" if c == len(hdrs) else "center")
                if c == 2 and isinstance(val, datetime.datetime):   # Data
                    cell.number_format = "dd/mm/yyyy hh:mm:ss"
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
        wb.save(path)
        QMessageBox.information(self, "Sucesso",
            f"{self._table.rowCount()} registros exportados para:\n{path}")

    def _exportar_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar como CSV", "", "CSV (*.csv)")
        if not path:
            return
        import csv
        hdrs = [COLS_DADOS[i] if i != 8 else "Valor"
                for i in range(len(COLS_DADOS))]
        linhas = []
        for row in self._linhas_visiveis(tipado=True):
            nova = []
            for val in row:
                if isinstance(val, datetime.datetime):
                    nova.append(val.strftime("%d/%m/%Y %H:%M:%S"))
                else:
                    nova.append(val)
            linhas.append(nova)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(hdrs)
            writer.writerows(linhas)
        QMessageBox.information(self, "Sucesso",
            f"{self._table.rowCount()} registros exportados para:\n{path}")

    def refresh(self):
        self._carregar()


# ═══════════════════════════════════════════════════════════
#  ABA IMPORTAÇÃO
# ═══════════════════════════════════════════════════════════

class AbaImport(QWidget):
    def __init__(self, aba_form: AbaForm):
        super().__init__()
        self._aba_form = aba_form
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 16, 12, 8)
        root.setSpacing(10)

        if not PANDAS_OK:
            root.addWidget(QLabel(
                "pandas não instalado.\nExecute: pip install pandas openpyxl xlrd"))
            return

        # ── título centralizado ───────────────────────────
        title = QLabel("Importar Arquivo")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size:16px; font-weight:bold; margin-bottom:6px;")
        root.addWidget(title)

        # ── container centralizado com formulário ─────────
        center_container = QWidget()
        center_layout = QHBoxLayout(center_container)
        center_layout.setContentsMargins(0, 0, 0, 0)

        form_widget = QWidget()
        form_widget.setMaximumWidth(500)
        form_layout = QFormLayout(form_widget)
        form_layout.setSpacing(10)
        form_layout.setContentsMargins(0, 0, 0, 0)
        form_layout.setLabelAlignment(Qt.AlignRight)

        # Arquivo
        arq_lbl = QLabel("Arquivo:")
        arq_row = QHBoxLayout()
        self._path = QLineEdit()
        self._path.setReadOnly(True)
        self._path.setMinimumWidth(260)
        btn_browse = QPushButton("...")
        btn_browse.setFixedWidth(36)
        btn_browse.clicked.connect(self._browse)
        arq_row.addWidget(self._path)
        arq_row.addWidget(btn_browse)
        form_layout.addRow(arq_lbl, arq_row)

        # Separador CSV
        sep_lbl = QLabel("Separador CSV:")
        self._sep = QLineEdit(";")
        self._sep.setFixedWidth(40)
        form_layout.addRow(sep_lbl, self._sep)

        center_layout.addStretch()
        center_layout.addWidget(form_widget)
        center_layout.addStretch()
        root.addWidget(center_container)

        # ── modo de importação (label + radio buttons centrados) ──
        modo_lbl = QLabel("Modo de importação:")
        modo_lbl.setAlignment(Qt.AlignCenter)
        modo_lbl.setStyleSheet("font-weight:bold; font-size:15px; margin-top:4px;")
        root.addWidget(modo_lbl)

        radio_container = QWidget()
        radio_layout = QHBoxLayout(radio_container)
        radio_layout.setContentsMargins(0, 0, 0, 0)
        self._rb_add = QRadioButton("Acrescentar ao banco existente")
        self._rb_ow  = QRadioButton("Sobrescrever banco (apaga tudo antes)")
        self._rb_ow.setStyleSheet("color:#c62828; font-size:15px; font-weight:bold")
        self._rb_add.setChecked(True)
        radio_layout.addStretch()
        radio_layout.addWidget(self._rb_add)
        radio_layout.addSpacing(20)
        radio_layout.addWidget(self._rb_ow)
        radio_layout.addStretch()
        root.addWidget(radio_container)

        # ── botão Importar centralizado ───────────────────
        btn_container = QWidget()
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 4, 0, 4)
        btn_imp = QPushButton("Importar")
        btn_imp.setStyleSheet(
            "QPushButton{background:#4CAF50;color:white;border-radius:6px;"
            "padding:10px 28px;font-weight:bold;font-size:14px;}"
            "QPushButton:hover{background:#43A047;border:1px solid rgba(0,0,0,0.2);}"
        )
        btn_imp.setFixedWidth(180)
        btn_imp.clicked.connect(self._importar)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_imp)
        btn_layout.addStretch()
        root.addWidget(btn_container)

        # ── status centralizado ───────────────────────────
        self._status = QLabel("")
        self._status.setAlignment(Qt.AlignCenter)
        root.addWidget(self._status)

        # ── preview (preenche espaço restante) ────────────
        grp_prev = QGroupBox("Pré-visualização (20 primeiras linhas)")
        lay_prev = QVBoxLayout(grp_prev)
        self._preview = QTableWidget(0, 0)
        self._preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._preview.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._preview.verticalHeader().setVisible(False)
        lay_prev.addWidget(self._preview)
        root.addWidget(grp_prev, 1)

    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Selecionar arquivo", "",
            "Planilhas (*.csv *.xls *.xlsx);;Todos (*.*)")
        if path:
            self._path.setText(path)
            self._load_preview(path)

    def _ler(self, path):
        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            sep = self._sep.text() or ";"
            try:
                return pd.read_csv(path, sep=sep, encoding="utf-8-sig", dtype=str)
            except UnicodeDecodeError:
                return pd.read_csv(path, sep=sep, encoding="latin1", dtype=str)
        elif ext == ".xlsx":
            return pd.read_excel(path, dtype=str, engine="openpyxl")
        elif ext == ".xls":
            return pd.read_excel(path, dtype=str, engine="xlrd")
        raise ValueError(f"Formato não suportado: {ext}")

    def _load_preview(self, path):
        try:
            df = self._ler(path)
            self._preview.clear()
            self._preview.setColumnCount(len(df.columns))
            self._preview.setHorizontalHeaderLabels(list(df.columns))
            self._preview.setRowCount(min(20, len(df)))
            for i, (_, row) in enumerate(df.head(20).iterrows()):
                for j, v in enumerate(row):
                    it = QTableWidgetItem(str(v))
                    self._preview.setItem(i, j, it)
            self._status.setText(f"{len(df)} linhas encontradas.")
            self._status.setStyleSheet("color:#1b5e20")
        except Exception as e:
            self._status.setText(f"Erro: {e}")
            self._status.setStyleSheet("color:#c62828")

    def _importar(self):
        path = self._path.text().strip()
        if not path:
            QMessageBox.warning(self, "Atenção", "Selecione um arquivo primeiro.")
            return
        modo = "sobrescrever" if self._rb_ow.isChecked() else "acrescentar"
        if modo == "sobrescrever":
            if QMessageBox.question(self, "Confirmar",
                    "Isso apagará TODOS os registros existentes. Continuar?",
                    QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return
        try:
            df = self._ler(path)
            importar_df(df, modo)
            self._status.setText(f"{len(df)} registros importados com sucesso.")
            self._status.setStyleSheet("color:#1b5e20")
            self._aba_form.refresh()
        except Exception as e:
            self._status.setText(f"Erro: {e}")
            self._status.setStyleSheet("color:#c62828")


# ═══════════════════════════════════════════════════════════
#  ABA CATEGORIAS (gerenciamento de Categoria/Sub-Categoria/Transação)
# ═══════════════════════════════════════════════════════════

class AbaCategorias(QWidget):
    """Cadastro de Categorias, Sub-Categorias (vinculadas a uma Categoria) e
    Transações, usados como listas fechadas nos combos da aba Dados."""

    def __init__(self, aba_form: AbaForm):
        super().__init__()
        self._aba_form = aba_form
        self._build()
        self._recarregar_categorias()
        self._recarregar_transacoes()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)

        info = QLabel(
            "Cadastre aqui as Categorias, Sub-Categorias e Transações que poderão ser "
            "escolhidas na aba \"Dados\". Cada Sub-Categoria pertence a uma Categoria.")
        info.setStyleSheet("color:#555; font-size:11px;")
        info.setWordWrap(True)
        root.addWidget(info)

        cols = QHBoxLayout()
        cols.setSpacing(12)

        # ── Categorias ─────────────────────────────────────
        grp_cat = QGroupBox("Categorias")
        lay_cat = QVBoxLayout(grp_cat)
        self._lst_cat = QListWidget()
        self._lst_cat.currentItemChanged.connect(self._on_categoria_selecionada)
        lay_cat.addWidget(self._lst_cat)
        row_cat = QHBoxLayout()
        row_cat.addWidget(_btn("Adicionar", "#4CAF50", self._add_categoria, 90))
        row_cat.addWidget(_btn("Renomear",  "#1565C0", self._ren_categoria, 90))
        row_cat.addWidget(_btn("Excluir",   "#f44336", self._del_categoria, 90))
        lay_cat.addLayout(row_cat)
        cols.addWidget(grp_cat, 1)

        # ── Sub-Categorias ──────────────────────────────────
        grp_sub = QGroupBox("Sub-Categorias")
        lay_sub = QVBoxLayout(grp_sub)
        self._lbl_sub = QLabel("Selecione uma categoria à esquerda")
        self._lbl_sub.setStyleSheet("color:#777; font-size:11px;")
        lay_sub.addWidget(self._lbl_sub)
        self._lst_sub = QListWidget()
        lay_sub.addWidget(self._lst_sub)
        row_sub = QHBoxLayout()
        row_sub.addWidget(_btn("Adicionar", "#4CAF50", self._add_subcategoria, 90))
        row_sub.addWidget(_btn("Renomear",  "#1565C0", self._ren_subcategoria, 90))
        row_sub.addWidget(_btn("Excluir",   "#f44336", self._del_subcategoria, 90))
        lay_sub.addLayout(row_sub)
        cols.addWidget(grp_sub, 1)

        # ── Transações ──────────────────────────────────────
        grp_tran = QGroupBox("Transações")
        lay_tran = QVBoxLayout(grp_tran)
        self._lst_tran = QListWidget()
        lay_tran.addWidget(self._lst_tran)
        row_tran = QHBoxLayout()
        row_tran.addWidget(_btn("Adicionar", "#4CAF50", self._add_transacao, 90))
        row_tran.addWidget(_btn("Renomear",  "#1565C0", self._ren_transacao, 90))
        row_tran.addWidget(_btn("Excluir",   "#f44336", self._del_transacao, 90))
        lay_tran.addLayout(row_tran)
        cols.addWidget(grp_tran, 1)

        root.addLayout(cols, 1)

    # ── recarregar listas ─────────────────────────────────
    def recarregar(self):
        """Recarrega todas as listas (ex.: após importação em outra aba)."""
        self._recarregar_categorias()
        self._recarregar_transacoes()

    def _recarregar_categorias(self):
        atual = self._lst_cat.currentItem().text() if self._lst_cat.currentItem() else None
        self._lst_cat.clear()
        for cid, nome in listar_categorias():
            item = QListWidgetItem(nome)
            item.setData(Qt.UserRole, cid)
            self._lst_cat.addItem(item)
        if atual:
            for i in range(self._lst_cat.count()):
                if self._lst_cat.item(i).text() == atual:
                    self._lst_cat.setCurrentRow(i)
                    return
        self._recarregar_subcategorias()

    def _recarregar_subcategorias(self):
        item = self._lst_cat.currentItem()
        self._lst_sub.clear()
        if not item:
            self._lbl_sub.setText("Selecione uma categoria à esquerda")
            return
        self._lbl_sub.setText(f"Sub-categorias de: {item.text()}")
        cid = item.data(Qt.UserRole)
        for sid, nome in listar_subcategorias(cid):
            sub_item = QListWidgetItem(nome)
            sub_item.setData(Qt.UserRole, sid)
            self._lst_sub.addItem(sub_item)

    def _recarregar_transacoes(self):
        self._lst_tran.clear()
        for tid, nome in listar_transacoes():
            item = QListWidgetItem(nome)
            item.setData(Qt.UserRole, tid)
            self._lst_tran.addItem(item)

    def _on_categoria_selecionada(self, *_):
        self._recarregar_subcategorias()

    def _avisar_form(self):
        """Atualiza a aba Dados (combos e tabela) após qualquer alteração de cadastro."""
        self._aba_form.refresh()

    # ── Categorias ─────────────────────────────────────────
    def _add_categoria(self):
        nome, ok = QInputDialog.getText(self, "Nova Categoria", "Nome da categoria:")
        if not ok or not nome.strip():
            return
        ok, msg = inserir_categoria(nome)
        if not ok:
            QMessageBox.warning(self, "Atenção", msg)
            return
        self._recarregar_categorias()
        self._avisar_form()

    def _ren_categoria(self):
        item = self._lst_cat.currentItem()
        if not item:
            QMessageBox.information(self, "Info", "Selecione uma categoria.")
            return
        novo, ok = QInputDialog.getText(self, "Renomear Categoria", "Novo nome:",
                                         text=item.text())
        if not ok or not novo.strip():
            return
        ok, msg = renomear_categoria(item.data(Qt.UserRole), novo)
        if not ok:
            QMessageBox.warning(self, "Atenção", msg)
            return
        self._recarregar_categorias()
        self._avisar_form()

    def _del_categoria(self):
        item = self._lst_cat.currentItem()
        if not item:
            QMessageBox.information(self, "Info", "Selecione uma categoria.")
            return
        if QMessageBox.question(self, "Confirmar",
                f"Excluir a categoria \"{item.text()}\"?",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        ok, msg = excluir_categoria(item.data(Qt.UserRole))
        if not ok:
            QMessageBox.warning(self, "Não foi possível excluir", msg)
            return
        self._recarregar_categorias()
        self._avisar_form()

    # ── Sub-Categorias ──────────────────────────────────────
    def _categoria_atual_id(self):
        item = self._lst_cat.currentItem()
        return item.data(Qt.UserRole) if item else None

    def _add_subcategoria(self):
        cid = self._categoria_atual_id()
        if cid is None:
            QMessageBox.information(self, "Info", "Selecione primeiro uma categoria.")
            return
        nome, ok = QInputDialog.getText(self, "Nova Sub-Categoria", "Nome da sub-categoria:")
        if not ok or not nome.strip():
            return
        ok, msg = inserir_subcategoria(nome, cid)
        if not ok:
            QMessageBox.warning(self, "Atenção", msg)
            return
        self._recarregar_subcategorias()
        self._avisar_form()

    def _ren_subcategoria(self):
        item = self._lst_sub.currentItem()
        if not item:
            QMessageBox.information(self, "Info", "Selecione uma sub-categoria.")
            return
        novo, ok = QInputDialog.getText(self, "Renomear Sub-Categoria", "Novo nome:",
                                         text=item.text())
        if not ok or not novo.strip():
            return
        ok, msg = renomear_subcategoria(item.data(Qt.UserRole), novo)
        if not ok:
            QMessageBox.warning(self, "Atenção", msg)
            return
        self._recarregar_subcategorias()
        self._avisar_form()

    def _del_subcategoria(self):
        item = self._lst_sub.currentItem()
        if not item:
            QMessageBox.information(self, "Info", "Selecione uma sub-categoria.")
            return
        if QMessageBox.question(self, "Confirmar",
                f"Excluir a sub-categoria \"{item.text()}\"?",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        ok, msg = excluir_subcategoria(item.data(Qt.UserRole))
        if not ok:
            QMessageBox.warning(self, "Não foi possível excluir", msg)
            return
        self._recarregar_subcategorias()
        self._avisar_form()

    # ── Transações ───────────────────────────────────────────
    def _add_transacao(self):
        nome, ok = QInputDialog.getText(self, "Nova Transação", "Nome da transação:")
        if not ok or not nome.strip():
            return
        ok, msg = inserir_transacao(nome)
        if not ok:
            QMessageBox.warning(self, "Atenção", msg)
            return
        self._recarregar_transacoes()
        self._avisar_form()

    def _ren_transacao(self):
        item = self._lst_tran.currentItem()
        if not item:
            QMessageBox.information(self, "Info", "Selecione uma transação.")
            return
        novo, ok = QInputDialog.getText(self, "Renomear Transação", "Novo nome:",
                                         text=item.text())
        if not ok or not novo.strip():
            return
        ok, msg = renomear_transacao(item.data(Qt.UserRole), novo)
        if not ok:
            QMessageBox.warning(self, "Atenção", msg)
            return
        self._recarregar_transacoes()
        self._avisar_form()

    def _del_transacao(self):
        item = self._lst_tran.currentItem()
        if not item:
            QMessageBox.information(self, "Info", "Selecione uma transação.")
            return
        if QMessageBox.question(self, "Confirmar",
                f"Excluir a transação \"{item.text()}\"?",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        ok, msg = excluir_transacao(item.data(Qt.UserRole))
        if not ok:
            QMessageBox.warning(self, "Não foi possível excluir", msg)
            return
        self._recarregar_transacoes()
        self._avisar_form()


# ═══════════════════════════════════════════════════════════
#  ABA TABELA DINÂMICA
# ═══════════════════════════════════════════════════════════

ALL_DIM = ["Ano", "Mes", "Categoria", "Sub_Categoria", "Transacao", "Descricao"]

class AbaPivot(QWidget):
    def __init__(self):
        super().__init__()
        self._export_rows = []
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 8, 12, 8)
        root.setSpacing(6)

        # ── estrutura ─────────────────────────────────────
        grp_str = QGroupBox("Estrutura")
        lay_str = QGridLayout(grp_str)
        lay_str.setSpacing(6)

        def lbl(t): return QLabel(t)
        def combo(vals, default):
            cb = QComboBox(); cb.addItems(vals)
            cb.setCurrentText(default); return cb

        lay_str.addWidget(lbl("Linha 1 (grupo):"),    0, 0, Qt.AlignRight)
        self._row1 = combo(ALL_DIM, "Categoria")
        lay_str.addWidget(self._row1, 0, 1)
        self._btn_excl1 = QPushButton("Excluir itens ▼")
        self._btn_excl1.setFixedWidth(130)
        self._btn_excl1.clicked.connect(lambda: self._abrir_menu_exclusao(1))
        lay_str.addWidget(self._btn_excl1, 0, 2)

        lay_str.addWidget(lbl("Linha 2 (subgrupo):"), 0, 3, Qt.AlignRight)
        self._row2 = combo(["(nenhuma)"] + ALL_DIM, "Descricao")
        lay_str.addWidget(self._row2, 0, 4)
        self._btn_excl2 = QPushButton("Excluir itens ▼")
        self._btn_excl2.setFixedWidth(130)
        self._btn_excl2.clicked.connect(lambda: self._abrir_menu_exclusao(2))
        lay_str.addWidget(self._btn_excl2, 0, 5)

        lay_str.addWidget(lbl("Colunas:"),            0, 6, Qt.AlignRight)
        self._cols = combo(["(nenhuma)"] + ALL_DIM, "Mes")
        lay_str.addWidget(self._cols, 0, 7)

        lay_str.addWidget(lbl("Agregar:"),            1, 0, Qt.AlignRight)
        self._agg = combo(["sum", "count", "mean", "min", "max"], "sum")
        self._agg.setFixedWidth(90)
        lay_str.addWidget(self._agg, 1, 1)

        self._chk_sub = QCheckBox("Subtotais"); self._chk_sub.setChecked(True)
        lay_str.addWidget(self._chk_sub, 1, 3, 1, 2)
        self._chk_total = QCheckBox("Total Geral"); self._chk_total.setChecked(True)
        lay_str.addWidget(self._chk_total, 1, 5, 1, 2)
        self._chk_pct = QCheckBox("Mostrar como %"); self._chk_pct.setChecked(False)
        lay_str.addWidget(self._chk_pct, 1, 7)
        root.addWidget(grp_str)

        # sets de exclusão e estado de expansão (populados dinamicamente)
        self._excluidos1 = set()
        self._excluidos2 = set()
        self._expandidos = set()  # nomes dos grupos expandidos

        # ── filtros ───────────────────────────────────────
        grp_flt = QGroupBox("Filtros de Relatório")
        lay_flt = QGridLayout(grp_flt)
        lay_flt.setSpacing(6)

        lay_flt.addWidget(lbl("Ano:"),           0, 0, Qt.AlignRight)
        self._f_ano  = QComboBox(); self._f_ano.setFixedWidth(80)
        lay_flt.addWidget(self._f_ano, 0, 1)

        lay_flt.addWidget(lbl("Mês:"),           0, 2, Qt.AlignRight)
        self._f_mes  = QComboBox(); self._f_mes.setMinimumWidth(150)
        lay_flt.addWidget(self._f_mes, 0, 3)

        lay_flt.addWidget(lbl("Categoria:"),     0, 4, Qt.AlignRight)
        self._f_cat  = QComboBox(); self._f_cat.setMinimumWidth(160)
        lay_flt.addWidget(self._f_cat, 0, 5)

        lay_flt.addWidget(lbl("Transação:"),     1, 0, Qt.AlignRight)
        self._f_tran = QComboBox(); self._f_tran.setMinimumWidth(200)
        lay_flt.addWidget(self._f_tran, 1, 1, 1, 2)

        lay_flt.addWidget(lbl("Sub-Categoria:"), 1, 4, Qt.AlignRight)
        self._f_sub  = QComboBox(); self._f_sub.setMinimumWidth(160)
        lay_flt.addWidget(self._f_sub, 1, 5)

        # radio buttons: todos / somente positivos / somente negativos
        lay_flt.addWidget(lbl("Valores:"), 2, 0, Qt.AlignRight)
        self._rb_todos = QRadioButton("Todos");          self._rb_todos.setChecked(True)
        self._rb_pos   = QRadioButton("Somente positivos")
        self._rb_neg   = QRadioButton("Somente negativos")
        self._rb_grp   = QButtonGroup(self)
        for rb in (self._rb_todos, self._rb_pos, self._rb_neg):
            self._rb_grp.addButton(rb)
        rb_row = QHBoxLayout()
        rb_row.setSpacing(16)
        for rb in (self._rb_todos, self._rb_pos, self._rb_neg):
            rb_row.addWidget(rb)
        rb_row.addStretch()
        lay_flt.addLayout(rb_row, 2, 1, 1, 5)

        # filtro por período (faixa de datas)
        self._chk_periodo = QCheckBox("Filtrar por período:")
        lay_flt.addWidget(self._chk_periodo, 3, 0, Qt.AlignRight)
        per_row = QHBoxLayout()
        per_row.setSpacing(6)
        per_row.addWidget(lbl("De:"))
        self._dt_de = QDateEdit(); self._dt_de.setCalendarPopup(True)
        self._dt_de.setDisplayFormat("dd/MM/yyyy")
        self._dt_de.setDate(QDate.currentDate())
        self._dt_de.setEnabled(False)
        per_row.addWidget(self._dt_de)
        per_row.addWidget(lbl("Até:"))
        self._dt_ate = QDateEdit(); self._dt_ate.setCalendarPopup(True)
        self._dt_ate.setDisplayFormat("dd/MM/yyyy")
        self._dt_ate.setDate(QDate.currentDate())
        self._dt_ate.setEnabled(False)
        per_row.addWidget(self._dt_ate)
        per_row.addStretch()
        lay_flt.addLayout(per_row, 3, 1, 1, 5)
        self._chk_periodo.toggled.connect(self._on_toggle_periodo)
        root.addWidget(grp_flt)

        # ── botões expandir/recolher ──────────────────────
        btn_tree_row = QHBoxLayout()
        btn_exp = QPushButton("▼  Expandir Tudo")
        btn_rec = QPushButton("▶  Recolher Tudo")
        for b in (btn_exp, btn_rec):
            b.setFixedHeight(28)
            b.setFixedWidth(150)
        btn_exp.clicked.connect(lambda: self._tree.expandAll())
        btn_rec.clicked.connect(lambda: self._tree.collapseAll())
        btn_tree_row.addWidget(btn_exp)
        btn_tree_row.addWidget(btn_rec)
        btn_tree_row.addStretch()
        root.addLayout(btn_tree_row)

        # ── resultado ─────────────────────────────────────
        self._tree = QTreeWidget()
        self._tree.setAlternatingRowColors(False)
        self._tree.setUniformRowHeights(True)
        self._tree.setRootIsDecorated(True)
        self._tree.setSortingEnabled(False)
        self._tree.header().setSectionResizeMode(QHeaderView.Interactive)
        self._tree.setFont(QFont("Segoe UI", 11))
        # ordenação por clique no cabeçalho (manual, sobre os dados)
        self._sort_col  = None    # None = ordem padrão (alfabética por grupo)
        self._sort_desc = True
        hdr = self._tree.header()
        hdr.setSectionsClickable(True)
        hdr.setSortIndicatorShown(True)
        hdr.sectionClicked.connect(self._on_header_click)
        root.addWidget(self._tree, 1)

        self._status = QLabel("")
        self._status.setStyleSheet("color:#555")
        root.addWidget(self._status)

        # ── conectar sinais para auto-gerar ───────────────
        for cb in (self._row1, self._row2, self._cols, self._agg,
                   self._f_ano, self._f_mes, self._f_cat, self._f_tran, self._f_sub):
            cb.currentIndexChanged.connect(self._gerar)
        for chk in (self._chk_sub, self._chk_total, self._chk_pct):
            chk.stateChanged.connect(self._gerar)
        for rb in (self._rb_todos, self._rb_pos, self._rb_neg):
            rb.toggled.connect(self._gerar)
        self._chk_periodo.toggled.connect(self._gerar)
        self._dt_de.dateChanged.connect(self._gerar)
        self._dt_ate.dateChanged.connect(self._gerar)

        self._tree.itemExpanded.connect(
            lambda item: self._on_expansao(item, True))
        self._tree.itemCollapsed.connect(
            lambda item: self._on_expansao(item, False))

        self._atualizar_filtros()
        self._restaurar_config_pivot()

    # ── dados ─────────────────────────────────────────────
    def _carregar_df(self):
        if not PANDAS_OK:
            return None
        cols, rows = buscar_todos()
        if not rows:
            return pd.DataFrame(columns=["id","Data","Mes","Ano","Categoria",
                                         "Sub_Categoria","Transacao","Descricao","Valor"])
        df = pd.DataFrame(rows, columns=cols)
        df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce").fillna(0)
        df["Mes"]   = pd.to_numeric(df["Mes"],   errors="coerce")
        df["Ano"]   = pd.to_numeric(df["Ano"],   errors="coerce")
        df["_DataDT"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
        return df

    def _on_toggle_periodo(self, ligado):
        self._dt_de.setEnabled(ligado)
        self._dt_ate.setEnabled(ligado)
        if ligado:
            hoje = QDate.currentDate()
            self._dt_de.blockSignals(True);  self._dt_de.setDate(hoje);  self._dt_de.blockSignals(False)
            self._dt_ate.blockSignals(True); self._dt_ate.setDate(hoje); self._dt_ate.blockSignals(False)

    def _atualizar_filtros(self):
        df = self._carregar_df()
        todos = ["(todos)"]
        if df is None or df.empty:
            for cb in (self._f_ano, self._f_cat, self._f_tran, self._f_sub, self._f_mes):
                cb.blockSignals(True)
                cb.clear()
                cb.addItems(todos)
                cb.blockSignals(False)
            self._tree.clear()
            return
        anos  = todos + sorted(df["Ano"].dropna().unique().astype(int).astype(str).tolist())
        cats  = todos + sorted(df["Categoria"].dropna().unique().tolist())
        trans = todos + sorted(df["Transacao"].dropna().unique().tolist())
        subs  = todos + sorted(df["Sub_Categoria"].dropna().unique().tolist())
        meses = todos + [f"{i} – {NOMES_MESES[i]}" for i in range(1, 13)]
        for cb, vals in ((self._f_ano, anos), (self._f_cat, cats),
                         (self._f_tran, trans), (self._f_sub, subs),
                         (self._f_mes, meses)):
            cur = cb.currentText()
            cb.blockSignals(True)
            cb.clear(); cb.addItems(vals)
            idx = cb.findText(cur)
            if idx >= 0: cb.setCurrentIndex(idx)
            cb.blockSignals(False)

    # ── restaurar configuração salva ──────────────────────
    def _restaurar_config_pivot(self):
        cfg = cfg_load().get("pivot_config")
        if not cfg:
            return
        # block signals during restore to avoid multiple _gerar calls
        widgets = [self._row1, self._row2, self._cols, self._agg,
                   self._f_ano, self._f_mes, self._f_cat, self._f_tran, self._f_sub,
                   self._chk_sub, self._chk_total, self._chk_pct,
                   self._rb_todos, self._rb_pos, self._rb_neg]
        for w in widgets:
            w.blockSignals(True)
        try:
            for attr, key in [("_row1", "row1"), ("_row2", "row2"),
                               ("_cols", "cols"), ("_agg", "agg"),
                               ("_f_ano", "f_ano"), ("_f_mes", "f_mes"),
                               ("_f_cat", "f_cat"), ("_f_tran", "f_tran"),
                               ("_f_sub", "f_sub")]:
                val = cfg.get(key)
                if val is not None:
                    cb = getattr(self, attr)
                    idx = cb.findText(val)
                    if idx >= 0:
                        cb.setCurrentIndex(idx)
            if "subtotais" in cfg:
                self._chk_sub.setChecked(bool(cfg["subtotais"]))
            if "total_geral" in cfg:
                self._chk_total.setChecked(bool(cfg["total_geral"]))
            if "mostrar_pct" in cfg:
                self._chk_pct.setChecked(bool(cfg["mostrar_pct"]))
            fv = cfg.get("filtro_valor", "todos")
            if fv == "pos":  self._rb_pos.setChecked(True)
            elif fv == "neg": self._rb_neg.setChecked(True)
            else:             self._rb_todos.setChecked(True)
            if "excluidos1" in cfg:
                self._excluidos1 = set(cfg["excluidos1"])
            if "excluidos2" in cfg:
                self._excluidos2 = set(cfg["excluidos2"])
            if "expandidos" in cfg:
                self._expandidos = set(cfg["expandidos"])
            if cfg.get("sort_col") is not None:
                self._sort_col = int(cfg["sort_col"])
            self._sort_desc = bool(cfg.get("sort_desc", True))
        finally:
            for w in widgets:
                w.blockSignals(False)

    # ── ordenação por clique no cabeçalho ─────────────────
    def _on_header_click(self, col: int):
        if self._sort_col == col:
            self._sort_desc = not self._sort_desc   # alterna direção
        else:
            self._sort_col  = col
            self._sort_desc = True                  # nova coluna: maior→menor
        self._gerar()

    # ── estado de expansão dos grupos ────────────────────
    def _on_expansao(self, item: QTreeWidgetItem, expandido: bool):
        nome = item.text(0)
        if expandido:
            self._expandidos.add(nome)
        else:
            self._expandidos.discard(nome)
        cfg_save({"pivot_config": {
            **cfg_load().get("pivot_config", {}),
            "expandidos": list(self._expandidos),
        }})

    # ── menu de exclusão de itens ─────────────────────────
    def _abrir_menu_exclusao(self, linha: int):
        df = self._carregar_df()
        if df is None or df.empty:
            return
        campo = self._row1.currentText() if linha == 1 else self._row2.currentText()
        if campo == "(nenhuma)":
            return
        excluidos = self._excluidos1 if linha == 1 else self._excluidos2
        btn      = self._btn_excl1  if linha == 1 else self._btn_excl2

        itens = sorted(df[campo].dropna().unique().tolist(), key=str)

        menu = QMenu(self)
        menu.setStyleSheet("QMenu { padding:4px; } QCheckBox { padding:4px 8px; }")

        checks = []
        for item in itens:
            wa = QWidgetAction(menu)
            chk = QCheckBox(str(item), menu)
            chk.setChecked(str(item) in excluidos)
            item_str = str(item)
            def make_toggle(s, ex):
                def toggle(checked):
                    if checked: ex.add(s)
                    else:        ex.discard(s)
                    self._gerar()
                return toggle
            chk.toggled.connect(make_toggle(item_str, excluidos))
            wa.setDefaultWidget(chk)
            menu.addAction(wa)
            checks.append(chk)

        menu.addSeparator()
        act_todos = menu.addAction("Marcar todos")
        act_todos.triggered.connect(lambda: self._marcar_todos_exclusao(checks, True))
        act_nenhum = menu.addAction("Desmarcar todos")
        act_nenhum.triggered.connect(lambda: self._marcar_todos_exclusao(checks, False))

        menu.exec_(btn.mapToGlobal(btn.rect().bottomLeft()))

    def _marcar_todos_exclusao(self, checks, marcado: bool):
        for chk in checks:
            chk.setChecked(marcado)

    # ── gerar ─────────────────────────────────────────────
    def _gerar(self):
        self._atualizar_filtros()
        df = self._carregar_df()
        if df is None:
            QMessageBox.warning(self, "Erro", "pandas não instalado.")
            return
        if df.empty:
            return

        # filtros
        if self._f_ano.currentText()  != "(todos)":
            df = df[df["Ano"] == int(self._f_ano.currentText())]
        mes_sel = self._f_mes.currentText()
        if mes_sel != "(todos)":
            df = df[df["Mes"] == int(mes_sel.split(" – ")[0])]
        if self._f_cat.currentText()  != "(todos)":
            df = df[df["Categoria"]     == self._f_cat.currentText()]
        if self._f_tran.currentText() != "(todos)":
            df = df[df["Transacao"]     == self._f_tran.currentText()]
        if self._f_sub.currentText()  != "(todos)":
            df = df[df["Sub_Categoria"] == self._f_sub.currentText()]
        if self._chk_periodo.isChecked():
            de  = self._dt_de.date().toPyDate()
            ate = self._dt_ate.date().toPyDate()
            df = df[(df["_DataDT"].dt.date >= de) & (df["_DataDT"].dt.date <= ate)]

        # exclusões de itens de linha 1 e linha 2
        row1_campo = self._row1.currentText()
        row2_campo = self._row2.currentText()
        if self._excluidos1 and row1_campo in df.columns:
            df = df[~df[row1_campo].astype(str).isin(self._excluidos1)]
        if self._excluidos2 and row2_campo in df.columns:
            df = df[~df[row2_campo].astype(str).isin(self._excluidos2)]

        # filtro positivos / negativos
        if self._rb_pos.isChecked():
            df = df[df["Valor"] > 0]
        elif self._rb_neg.isChecked():
            df = df[df["Valor"] < 0]

        if df.empty:
            self._tree.clear()
            self._status.setText("Nenhum dado para os filtros selecionados.")
            return

        row1     = self._row1.currentText()
        row2     = self._row2.currentText()
        col_fld  = self._cols.currentText()
        agg      = self._agg.currentText()
        use_row2 = row2 != "(nenhuma)"
        use_cols = col_fld != "(nenhuma)"

        col_vals = (
            sorted(df[col_fld].dropna().unique().tolist(),
                   key=lambda x: (int(x) if str(x).lstrip("-").isdigit() else 0, str(x)))
            if use_cols else ["Valor"]
        )

        def agregar(sub):
            if sub.empty: return 0.0
            if agg == "sum":   return float(sub["Valor"].sum())
            if agg == "count": return float(sub["Valor"].count())
            v = getattr(sub["Valor"], agg)()
            return 0.0 if pd.isna(v) else float(v)

        def vals_por_col(sub):
            return ({str(cv): agregar(sub[sub[col_fld] == cv]) for cv in col_vals}
                    if use_cols else {"Valor": agregar(sub)})

        # ── configurar QTreeWidget ────────────────────────
        hdrs = [f"{row1}" + (f" / {row2}" if use_row2 else "")] + \
               [str(cv) for cv in col_vals] + ["Total Geral"]
        self._tree.clear()
        self._tree.setColumnCount(len(hdrs))
        self._tree.setHeaderLabels(hdrs)
        self._tree.header().setSectionResizeMode(0, QHeaderView.Interactive)
        self._tree.setColumnWidth(0, 210)
        for c in range(1, len(hdrs)):
            self._tree.setColumnWidth(c, 105)
            self._tree.header().setSectionResizeMode(c, QHeaderView.Interactive)

        # alinhamento do cabeçalho: rótulo à esquerda, colunas intermediárias
        # centralizadas, Total Geral à direita (mesmo alinhamento dos dados)
        hdr_item = self._tree.headerItem()
        hdr_item.setTextAlignment(0, Qt.AlignLeft | Qt.AlignVCenter)
        for c in range(1, len(hdrs) - 1):
            hdr_item.setTextAlignment(c, Qt.AlignCenter)
        hdr_item.setTextAlignment(len(hdrs) - 1, Qt.AlignRight | Qt.AlignVCenter)

        font_bold = QFont("Segoe UI", 11, QFont.Bold)
        font_reg  = QFont("Segoe UI", 11)

        usar_pct = self._chk_pct.isChecked()

        # ── 1ª passagem: calcular todos os valores e o total geral ──
        # os totais (linha/coluna/geral) são recalculados aplicando a MESMA
        # agregação sobre o conjunto de registros correspondente — somar os
        # subtotais já agregados só é correto para "sum"; para mean/min/max/
        # count isso daria um resultado errado.
        grupos = sorted(df[row1].dropna().unique().tolist())
        dados_grupos = []
        for g in grupos:
            g_df  = df[df[row1] == g]
            g_cv  = vals_por_col(g_df)
            g_tot = agregar(g_df)
            subgrupos_dados = []
            if use_row2:
                for sg in sorted(g_df[row2].dropna().unique().tolist()):
                    sg_df = g_df[g_df[row2] == sg]
                    sg_cv = vals_por_col(sg_df)
                    sg_tot = agregar(sg_df)
                    subgrupos_dados.append((sg, sg_cv, sg_tot))
            dados_grupos.append((g, g_cv, g_tot, subgrupos_dados))

        grand = ({str(cv): agregar(df[df[col_fld] == cv]) for cv in col_vals}
                 if use_cols else {"Valor": agregar(df)})
        gt_tot = agregar(df)

        # ── ordenação por coluna (sobre os dados, não a árvore) ──
        # reseta se a coluna salva não existe mais nesta configuração
        if self._sort_col is not None and self._sort_col >= len(hdrs):
            self._sort_col = None

        if self._sort_col is not None:
            sc = self._sort_col

            def chave_valor(g_cv, g_tot):
                """Valor numérico da coluna sc para ordenação."""
                if sc == 0:
                    return None
                if sc == len(hdrs) - 1:        # coluna Total Geral
                    return g_tot
                return g_cv.get(str(col_vals[sc - 1]), 0.0)

            if sc == 0:
                # coluna de rótulo: ordena pelo próprio valor do grupo
                dados_grupos.sort(key=lambda t: t[0], reverse=self._sort_desc)
            else:
                dados_grupos.sort(
                    key=lambda t: chave_valor(t[1], t[2]), reverse=self._sort_desc)

            # ordena também os subgrupos dentro de cada grupo
            for _g, _cv, _tot, subs in dados_grupos:
                if sc == 0:
                    subs.sort(key=lambda s: s[0], reverse=self._sort_desc)
                else:
                    subs.sort(key=lambda s: chave_valor(s[1], s[2]),
                              reverse=self._sort_desc)

        # indicador visual de ordenação no cabeçalho
        hdr = self._tree.header()
        hdr.blockSignals(True)
        if self._sort_col is None:
            hdr.setSortIndicatorShown(False)
        else:
            hdr.setSortIndicatorShown(True)
            hdr.setSortIndicator(
                self._sort_col,
                Qt.DescendingOrder if self._sort_desc else Qt.AscendingOrder)
        hdr.blockSignals(False)

        def fmt_cel(v, is_total_geral_cell=False):
            """Formata célula em valor ou %."""
            if not usar_pct:
                return fmt_valor(v)
            if is_total_geral_cell:
                return "100,00%"
            base = abs(gt_tot) if gt_tot != 0 else 1.0
            return f"{v / base * 100:.2f}%".replace(".", ",")

        def cor_cel(v):
            return cor_valor(v)

        # ── 2ª passagem: montar a árvore ─────────────────────
        export_rows = [hdrs]

        for g, g_cv, g_tot, subgrupos_dados in dados_grupos:
            g_strs = [fmt_cel(g_cv.get(str(cv), 0)) for cv in col_vals] + \
                     [fmt_cel(g_tot, usar_pct and g_tot == gt_tot)]
            parent = QTreeWidgetItem([str(g)] + g_strs)
            parent.setFont(0, font_bold)
            parent.setBackground(0, QBrush(BG_GRUPO))
            for c in range(1, len(hdrs)):
                parent.setFont(c, font_bold)
                parent.setBackground(c, QBrush(BG_GRUPO))
                v = g_cv.get(str(col_vals[c-1]), g_tot) if c < len(hdrs)-1 else g_tot
                parent.setForeground(c, QBrush(cor_cel(v)))
                parent.setTextAlignment(c, Qt.AlignRight | Qt.AlignVCenter
                                         if c == len(hdrs) - 1 else Qt.AlignCenter)
            export_rows.append([str(g)] + g_strs)

            for sg, sg_cv, sg_tot in subgrupos_dados:
                sg_strs = [fmt_cel(sg_cv.get(str(cv), 0)) for cv in col_vals] + \
                          [fmt_cel(sg_tot)]
                child = QTreeWidgetItem([str(sg)] + sg_strs)
                child.setFont(0, font_reg)
                for c in range(1, len(hdrs)):
                    child.setFont(c, font_reg)
                    v = sg_cv.get(str(col_vals[c-1]), sg_tot) if c < len(hdrs)-1 else sg_tot
                    child.setForeground(c, QBrush(cor_cel(v)))
                    child.setTextAlignment(c, Qt.AlignRight | Qt.AlignVCenter
                                            if c == len(hdrs) - 1 else Qt.AlignCenter)
                parent.addChild(child)
                export_rows.append(["  " + str(sg)] + sg_strs)

            self._tree.addTopLevelItem(parent)
            parent.setExpanded(str(g) in self._expandidos)

        # Total Geral
        if self._chk_total.isChecked():
            gt_strs = [fmt_cel(grand.get(str(cv), 0)) for cv in col_vals] + \
                      [fmt_cel(gt_tot, usar_pct)]
            total_item = QTreeWidgetItem(["Total Geral"] + gt_strs)
            total_item.setFont(0, font_bold)
            total_item.setBackground(0, QBrush(BG_TOTAL))
            for c in range(1, len(hdrs)):
                total_item.setFont(c, font_bold)
                total_item.setBackground(c, QBrush(BG_TOTAL))
                v = grand.get(str(col_vals[c-1]), gt_tot) if c < len(hdrs)-1 else gt_tot
                total_item.setForeground(c, QBrush(cor_cel(v)))
                total_item.setTextAlignment(c, Qt.AlignRight | Qt.AlignVCenter
                                             if c == len(hdrs) - 1 else Qt.AlignCenter)
            self._tree.addTopLevelItem(total_item)
            export_rows.append(["Total Geral"] + gt_strs)

        self._export_rows = export_rows
        # auto-ajuste da coluna de rótulo; demais via ResizeToContents
        self._tree.header().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        for c in range(1, len(hdrs)):
            self._tree.header().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self._status.setText(
            f"{len(grupos)} grupos  |  {len(df)} registros  |  agreg: {agg}"
            + ("  —  clique no ▶ para expandir" if use_row2 else ""))

        # ── salvar configuração em config.json ────────────
        cfg_save({"pivot_config": {
            "row1":        self._row1.currentText(),
            "row2":        self._row2.currentText(),
            "cols":        self._cols.currentText(),
            "agg":         self._agg.currentText(),
            "subtotais":   self._chk_sub.isChecked(),
            "total_geral": self._chk_total.isChecked(),
            "mostrar_pct": self._chk_pct.isChecked(),
            "filtro_valor": "pos" if self._rb_pos.isChecked() else "neg" if self._rb_neg.isChecked() else "todos",
            "f_ano":       self._f_ano.currentText(),
            "f_mes":       self._f_mes.currentText(),
            "f_cat":       self._f_cat.currentText(),
            "f_tran":      self._f_tran.currentText(),
            "f_sub":       self._f_sub.currentText(),
            "excluidos1":  list(self._excluidos1),
            "excluidos2":  list(self._excluidos2),
            "expandidos":  list(self._expandidos),
            "sort_col":    self._sort_col,
            "sort_desc":   self._sort_desc,
        }})

    # ── exportar ──────────────────────────────────────────
    def _exportar(self):
        if not self._export_rows:
            QMessageBox.information(self, "Info", "Gere a tabela antes de exportar.")
            return
        if not OPENPYXL_OK:
            QMessageBox.critical(self, "Erro",
                "openpyxl não instalado.\nExecute: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar como", "", "Excel (*.xlsx)")
        if not path:
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tabela Dinâmica"
        for r_idx, row in enumerate(self._export_rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal="right" if c_idx > 1 else "left")
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                elif "Total" in str(row[0]):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="E2EFDA")
                elif c_idx > 1:
                    try:
                        v = float(str(val).replace("R$","").replace(".","")
                                         .replace(",",".").strip())
                        cell.font = Font(color="C62828" if v < 0 else "1B5E20")
                    except (ValueError, AttributeError):
                        pass
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
        wb.save(path)
        QMessageBox.information(self, "Sucesso", f"Exportado para:\n{path}")


# ═══════════════════════════════════════════════════════════
#  JANELA PRINCIPAL
# ═══════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════
#  ABA DASHBOARD
# ═══════════════════════════════════════════════════════════

class AbaDashboard(QWidget):

    DIMS = {
        "Categoria":     "Categoria",
        "Sub-Categoria": "Sub_Categoria",
        "Transação":     "Transacao",
        "Descrição":     "Descricao",
        "Mês":           "Mes",
        "Ano":           "Ano",
    }
    DIMS_TEMPORAIS = {"Mês", "Ano"}
    # padrão: (dimensão, tipo)
    DEFAULTS = [
        ("Categoria",     "Saídas"),
        ("Sub-Categoria", "Saídas"),
        ("Mês",           "Saídas"),
    ]

    def __init__(self):
        super().__init__()
        self._cfg_aplicado = False
        self._build()
        self._restaurar_config()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 10, 14, 10)
        root.setSpacing(10)

        # ── filtros globais ───────────────────────────────
        flt = QHBoxLayout()
        flt.addWidget(QLabel("Ano:"))
        self._f_ano = QComboBox(); self._f_ano.setFixedWidth(90)
        self._f_ano.currentIndexChanged.connect(self._preencher)
        self._f_ano.currentIndexChanged.connect(self._salvar_config)
        flt.addWidget(self._f_ano)
        flt.addWidget(QLabel("  Mês:"))
        self._f_mes = QComboBox(); self._f_mes.setMinimumWidth(150)
        self._f_mes.currentIndexChanged.connect(self._preencher)
        self._f_mes.currentIndexChanged.connect(self._salvar_config)
        flt.addWidget(self._f_mes)

        self._chk_periodo = QCheckBox("  Filtrar por período:")
        self._chk_periodo.toggled.connect(self._on_toggle_periodo)
        self._chk_periodo.toggled.connect(self._salvar_config)
        flt.addWidget(self._chk_periodo)
        flt.addWidget(QLabel("De:"))
        self._dt_de = QDateEdit(); self._dt_de.setCalendarPopup(True)
        self._dt_de.setDisplayFormat("dd/MM/yyyy")
        self._dt_de.setDate(QDate.currentDate())
        self._dt_de.setEnabled(False)
        self._dt_de.dateChanged.connect(self._preencher)
        flt.addWidget(self._dt_de)
        flt.addWidget(QLabel("Até:"))
        self._dt_ate = QDateEdit(); self._dt_ate.setCalendarPopup(True)
        self._dt_ate.setDisplayFormat("dd/MM/yyyy")
        self._dt_ate.setDate(QDate.currentDate())
        self._dt_ate.setEnabled(False)
        self._dt_ate.dateChanged.connect(self._preencher)
        flt.addWidget(self._dt_ate)

        flt.addStretch()
        root.addLayout(flt)

        # ── cartões de resumo ─────────────────────────────
        cards_row = QHBoxLayout()
        cards_row.setSpacing(14)
        self._card_ent = self._make_card("Total Entradas", "#1b5e20")
        self._card_sai = self._make_card("Total Saídas",   "#c62828")
        self._card_sal = self._make_card("Saldo",          "#1565C0")
        self._card_qtd = self._make_card("Registros",      "#6A1B9A")
        for c in (self._card_ent, self._card_sai, self._card_sal, self._card_qtd):
            cards_row.addWidget(c)
        root.addLayout(cards_row)

        # ── 3 tabelas configuráveis ───────────────────────
        tabelas_row = QHBoxLayout()
        tabelas_row.setSpacing(14)
        self._paineis = []
        for dim_def, tipo_def in self.DEFAULTS:
            grp = QGroupBox()
            lay = QVBoxLayout(grp)
            lay.setSpacing(4)

            # linha de controles: combo dimensão + combo tipo
            ctrl = QHBoxLayout()
            cb_dim = QComboBox()
            cb_dim.addItems(list(self.DIMS.keys()))
            cb_dim.setCurrentText(dim_def)
            cb_tipo = QComboBox()
            cb_tipo.addItems(["Saídas", "Entradas"])
            cb_tipo.setCurrentText(tipo_def)
            cb_tipo.setFixedWidth(100)
            cb_tipo.setVisible(dim_def not in self.DIMS_TEMPORAIS)
            cb_dim.currentIndexChanged.connect(
                lambda _, c=cb_dim, ct=cb_tipo: (
                    ct.setVisible(c.currentText() not in self.DIMS_TEMPORAIS),
                    self._preencher(),
                    self._salvar_config()
                )
            )
            cb_tipo.currentIndexChanged.connect(self._preencher)
            cb_tipo.currentIndexChanged.connect(self._salvar_config)
            ctrl.addWidget(cb_dim, 1)
            ctrl.addWidget(cb_tipo)
            lay.addLayout(ctrl)

            tbl = self._make_table()
            lay.addWidget(tbl)
            tabelas_row.addWidget(grp)
            self._paineis.append((grp, cb_dim, cb_tipo, tbl))
        root.addLayout(tabelas_row, 1)

        # ── maior gasto único ─────────────────────────────
        self._lbl_maior = QLabel("")
        self._lbl_maior.setStyleSheet("font-size:12px; color:#555; padding:4px;")
        root.addWidget(self._lbl_maior)

    def _make_card(self, titulo, cor):
        frame = QFrame()
        frame.setFrameShape(QFrame.StyledPanel)
        frame.setStyleSheet(
            f"QFrame{{background:#f5f5f5;border:2px solid {cor};border-radius:8px;}}")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(16, 10, 16, 10)
        lbl_t = QLabel(titulo)
        lbl_t.setStyleSheet(f"font-size:11px;font-weight:bold;color:{cor};border:none;")
        lbl_v = QLabel("—")
        lbl_v.setStyleSheet(f"font-size:20px;font-weight:bold;color:{cor};border:none;")
        lbl_v.setAlignment(Qt.AlignCenter)
        lay.addWidget(lbl_t)
        lay.addWidget(lbl_v)
        frame._lbl = lbl_v
        return frame

    def _make_table(self):
        t = QTableWidget(0, 3)
        t.setHorizontalHeaderLabels(["", "Total", "%"])
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.verticalHeader().setVisible(False)
        t.horizontalHeader().setStretchLastSection(False)
        t.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        t.setAlternatingRowColors(True)
        t.setFont(QFont("Segoe UI", 10))
        return t

    def _set_colunas_temporais(self, tbl, dim_nome):
        tbl.setColumnCount(4)
        tbl.setHorizontalHeaderLabels([dim_nome, "Entradas", "Saídas", "Saldo"])
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for c in (1, 2, 3):
            tbl.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)

    def _set_colunas_tipo(self, tbl, dim_nome):
        tbl.setColumnCount(3)
        tbl.setHorizontalHeaderLabels([dim_nome, "Total", "%"])
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)

    def _on_toggle_periodo(self, ligado):
        self._dt_de.setEnabled(ligado)
        self._dt_ate.setEnabled(ligado)
        if ligado:
            hoje = QDate.currentDate()
            self._dt_de.blockSignals(True);  self._dt_de.setDate(hoje);  self._dt_de.blockSignals(False)
            self._dt_ate.blockSignals(True); self._dt_ate.setDate(hoje); self._dt_ate.blockSignals(False)
        self._preencher()

    # ── persistência de configuração (config.json) ───────
    def _salvar_config(self):
        cfg_save({"dashboard_config": {
            "f_ano":          self._f_ano.currentText(),
            "f_mes":          self._f_mes.currentText(),
            "periodo_ligado": self._chk_periodo.isChecked(),
            "paineis": [
                {"dim": cb_dim.currentText(), "tipo": cb_tipo.currentText()}
                for _, cb_dim, cb_tipo, _ in self._paineis
            ],
        }})

    def _restaurar_config(self):
        cfg = cfg_load().get("dashboard_config")
        if not cfg:
            return
        self._cfg_pendente = cfg
        paineis_cfg = cfg.get("paineis", [])
        for (grp, cb_dim, cb_tipo, tbl), pcfg in zip(self._paineis, paineis_cfg):
            cb_dim.blockSignals(True); cb_tipo.blockSignals(True)
            dim = pcfg.get("dim")
            if dim and cb_dim.findText(dim) >= 0:
                cb_dim.setCurrentText(dim)
            tipo = pcfg.get("tipo")
            if tipo and cb_tipo.findText(tipo) >= 0:
                cb_tipo.setCurrentText(tipo)
            cb_tipo.setVisible(cb_dim.currentText() not in self.DIMS_TEMPORAIS)
            cb_dim.blockSignals(False); cb_tipo.blockSignals(False)
        if "periodo_ligado" in cfg:
            self._chk_periodo.blockSignals(True)
            self._chk_periodo.setChecked(bool(cfg["periodo_ligado"]))
            self._dt_de.setEnabled(self._chk_periodo.isChecked())
            self._dt_ate.setEnabled(self._chk_periodo.isChecked())
            self._chk_periodo.blockSignals(False)

    def atualizar(self):
        if not PANDAS_OK:
            return
        cols, rows = buscar_todos()
        if not rows:
            return
        df = pd.DataFrame(rows, columns=cols)
        df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce").fillna(0)
        df["Mes"]   = pd.to_numeric(df["Mes"],   errors="coerce")
        df["Ano"]   = pd.to_numeric(df["Ano"],   errors="coerce")
        df["_DataDT"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
        df = df[df["Ano"] != 1900]
        self._df_full = df

        anos  = ["(todos)"] + sorted(df["Ano"].dropna().unique().astype(int).astype(str).tolist())
        meses = ["(todos)"] + [f"{i} – {NOMES_MESES[i]}" for i in range(1, 13)]
        cfg_pendente = getattr(self, "_cfg_pendente", None) if not self._cfg_aplicado else None
        for cb, vals, cfg_key in ((self._f_ano, anos, "f_ano"), (self._f_mes, meses, "f_mes")):
            cur = cfg_pendente.get(cfg_key) if cfg_pendente else cb.currentText()
            cb.blockSignals(True)
            cb.clear(); cb.addItems(vals)
            idx = cb.findText(cur) if cur else -1
            cb.setCurrentIndex(idx if idx >= 0 else 0)
            cb.blockSignals(False)
        self._cfg_aplicado = True
        self._preencher()

    def _preencher(self):
        if not hasattr(self, "_df_full"):
            return
        df = self._df_full.copy()
        if self._f_ano.currentText() != "(todos)":
            df = df[df["Ano"] == int(self._f_ano.currentText())]
        mes_sel = self._f_mes.currentText()
        if mes_sel != "(todos)":
            df = df[df["Mes"] == int(mes_sel.split(" – ")[0])]
        if self._chk_periodo.isChecked():
            de  = self._dt_de.date().toPyDate()
            ate = self._dt_ate.date().toPyDate()
            df = df[(df["_DataDT"].dt.date >= de) & (df["_DataDT"].dt.date <= ate)]

        entradas = df[df["Valor"] > 0]["Valor"].sum()
        saidas   = df[df["Valor"] < 0]["Valor"].sum()
        saldo    = entradas + saidas

        self._card_ent._lbl.setText(fmt_valor(entradas))
        self._card_sai._lbl.setText(fmt_valor(saidas))
        self._card_qtd._lbl.setText(str(len(df)))
        cor_sal = "#1b5e20" if saldo >= 0 else "#c62828"
        self._card_sal.setStyleSheet(
            f"QFrame{{background:#f5f5f5;border:2px solid {cor_sal};border-radius:8px;}}")
        self._card_sal._lbl.setStyleSheet(
            f"font-size:20px;font-weight:bold;color:{cor_sal};border:none;")
        self._card_sal._lbl.setText(fmt_valor(saldo))

        for grp, cb_dim, cb_tipo, tbl in self._paineis:
            dim_nome = cb_dim.currentText()
            dim_col  = self.DIMS[dim_nome]
            if dim_nome in self.DIMS_TEMPORAIS:
                grp.setTitle(dim_nome)
                self._set_colunas_temporais(tbl, dim_nome)
                self._fill_tabela_temporal(tbl, df, dim_col, dim_nome)
            else:
                tipo = cb_tipo.currentText()
                grp.setTitle(f"{dim_nome}  —  {tipo}")
                self._set_colunas_tipo(tbl, dim_nome)
                self._fill_tabela(tbl, df, dim_col, dim_nome, tipo)

        # maior gasto único
        df_sai = df[df["Valor"] < 0]
        if not df_sai.empty:
            row = df_sai.loc[df_sai["Valor"].idxmin()]
            self._lbl_maior.setText(
                f"  Maior gasto único:  {row['Data']}  |  "
                f"{row['Categoria']}  |  {row['Descricao']}  |  "
                f"{fmt_valor(row['Valor'])}")
        else:
            self._lbl_maior.setText("")

    def _fill_tabela_temporal(self, tbl, df, col, dim_nome):
        tbl.setRowCount(0)
        grp = df.groupby(col)["Valor"].sum()
        chaves = sorted(grp.index.tolist())
        for chave in chaves:
            if str(chave).strip() == "":
                continue
            if col == "Mes":
                label = f"{int(chave)} – {NOMES_MESES.get(int(chave), '')}"
            else:
                label = str(int(chave))
            sub = df[df[col] == chave]
            ent = sub[sub["Valor"] > 0]["Valor"].sum()
            sai = sub[sub["Valor"] < 0]["Valor"].sum()
            sal = ent + sai
            r = tbl.rowCount(); tbl.insertRow(r)
            tbl.setItem(r, 0, QTableWidgetItem(label))
            for c_idx, val in enumerate((ent, sai, sal), start=1):
                it = QTableWidgetItem(fmt_valor(val))
                it.setForeground(QBrush(cor_valor(val)))
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tbl.setItem(r, c_idx, it)
        tbl.resizeColumnsToContents()
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)

    def _fill_tabela(self, tbl, df, col, dim_nome, tipo):
        tbl.setHorizontalHeaderLabels([dim_nome, "Total", "%"])
        tbl.setRowCount(0)

        # filtrar entradas ou saídas
        if tipo == "Saídas":
            df_f = df[df["Valor"] < 0]
            base = abs(df_f["Valor"].sum()) if not df_f.empty else 1.0
        else:
            df_f = df[df["Valor"] > 0]
            base = df_f["Valor"].sum() if not df_f.empty else 1.0
        if base == 0:
            base = 1.0

        # agrupar e ordenar pelo valor absoluto (maior primeiro)
        grp = df_f.groupby(col)["Valor"].sum()
        if tipo == "Saídas":
            grp = grp.sort_values()          # mais negativo primeiro
        else:
            grp = grp.sort_values(ascending=False)  # maior entrada primeiro

        for chave, val in grp.items():
            if str(chave).strip() == "":
                continue
            if col == "Mes":
                label = f"{int(chave)} – {NOMES_MESES.get(int(chave), '')}"
            else:
                label = str(chave)

            pct = abs(val) / base * 100
            r = tbl.rowCount(); tbl.insertRow(r)
            tbl.setItem(r, 0, QTableWidgetItem(label))
            it_val = QTableWidgetItem(fmt_valor(val))
            it_val.setForeground(QBrush(cor_valor(val)))
            it_val.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            tbl.setItem(r, 1, it_val)
            it_pct = QTableWidgetItem(f"{pct:.1f}%")
            it_pct.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            tbl.setItem(r, 2, it_pct)
        tbl.resizeColumnsToContents()


# ═══════════════════════════════════════════════════════════
#  ABA TENDÊNCIAS
# ═══════════════════════════════════════════════════════════

class AbaTendencias(QWidget):

    DIMS = {
        "Categoria":     "Categoria",
        "Sub-Categoria": "Sub_Categoria",
        "Transação":     "Transacao",
    }

    def __init__(self):
        super().__init__()
        self._cfg_aplicado = False
        self._build()
        self._restaurar_config()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 10, 14, 10)
        root.setSpacing(10)

        # ── filtros globais ───────────────────────────────
        flt = QHBoxLayout()
        flt.addWidget(QLabel("Ano:"))
        self._f_ano = QComboBox(); self._f_ano.setFixedWidth(90)
        self._f_ano.currentIndexChanged.connect(self._preencher)
        self._f_ano.currentIndexChanged.connect(self._salvar_config)
        flt.addWidget(self._f_ano)
        flt.addWidget(QLabel("  Mês:"))
        self._f_mes = QComboBox(); self._f_mes.setMinimumWidth(150)
        self._f_mes.currentIndexChanged.connect(self._preencher)
        self._f_mes.currentIndexChanged.connect(self._salvar_config)
        flt.addWidget(self._f_mes)

        self._chk_periodo = QCheckBox("  Filtrar por período:")
        self._chk_periodo.toggled.connect(self._on_toggle_periodo)
        self._chk_periodo.toggled.connect(self._salvar_config)
        flt.addWidget(self._chk_periodo)
        flt.addWidget(QLabel("De:"))
        self._dt_de = QDateEdit(); self._dt_de.setCalendarPopup(True)
        self._dt_de.setDisplayFormat("dd/MM/yyyy")
        self._dt_de.setDate(QDate.currentDate())
        self._dt_de.setEnabled(False)
        self._dt_de.dateChanged.connect(self._preencher)
        flt.addWidget(self._dt_de)
        flt.addWidget(QLabel("Até:"))
        self._dt_ate = QDateEdit(); self._dt_ate.setCalendarPopup(True)
        self._dt_ate.setDisplayFormat("dd/MM/yyyy")
        self._dt_ate.setDate(QDate.currentDate())
        self._dt_ate.setEnabled(False)
        self._dt_ate.dateChanged.connect(self._preencher)
        flt.addWidget(self._dt_ate)

        flt.addWidget(QLabel("  Agrupar por:"))
        self._cb_gran = QComboBox(); self._cb_gran.setFixedWidth(80)
        self._cb_gran.addItems(["Mês", "Ano"])
        self._cb_gran.currentIndexChanged.connect(self._preencher)
        self._cb_gran.currentIndexChanged.connect(self._salvar_config)
        flt.addWidget(self._cb_gran)

        flt.addStretch()
        root.addLayout(flt)

        # ── card de destaque: tendência do saldo ──────────
        self._card_tend_saldo = self._make_card_destaque("Tendência do Saldo")
        root.addWidget(self._card_tend_saldo)

        # ── 3 painéis: Categoria / Sub-Categoria / Transação
        paineis_row = QHBoxLayout()
        paineis_row.setSpacing(14)
        self._paineis = []
        for nome_dim, col_dim in self.DIMS.items():
            grp = QGroupBox(nome_dim)
            lay = QVBoxLayout(grp)
            lay.setSpacing(6)

            lst = QListWidget()
            ctrl = QHBoxLayout()
            ctrl.addStretch()
            ctrl.addWidget(_btn("Marcar todos", "#1565C0",
                                  lambda _, l=lst: self._marcar_todos(l), 105))
            ctrl.addWidget(_btn("Nenhum", "#757575",
                                  lambda _, l=lst: self._desmarcar_todos(l), 80))
            lay.addLayout(ctrl)

            lst.itemChanged.connect(self._preencher)
            lay.addWidget(lst, 1)

            card_media = self._make_card_painel()
            lay.addWidget(card_media)

            paineis_row.addWidget(grp)
            self._paineis.append({"col": col_dim, "lst": lst, "card": card_media})
        root.addLayout(paineis_row, 1)

    def _make_card_destaque(self, titulo):
        frame = QFrame()
        frame.setFrameShape(QFrame.StyledPanel)
        frame.setStyleSheet(
            "QFrame{background:#f5f5f5;border:2px solid #1565C0;border-radius:8px;}")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(16, 10, 16, 10)
        lbl_t = QLabel(titulo)
        lbl_t.setStyleSheet("font-size:12px;font-weight:bold;color:#1565C0;border:none;")
        lbl_v = QLabel("—")
        lbl_v.setStyleSheet("font-size:22px;font-weight:bold;color:#1565C0;border:none;")
        lbl_v.setAlignment(Qt.AlignCenter)
        lay.addWidget(lbl_t)
        lay.addWidget(lbl_v)
        frame._lbl = lbl_v
        return frame

    def _make_card_painel(self):
        frame = QFrame()
        frame.setFrameShape(QFrame.StyledPanel)
        frame.setStyleSheet(
            "QFrame{background:#f5f5f5;border:1px solid #cccccc;border-radius:6px;}")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(2)
        lbl_t = QLabel("Média por período")
        lbl_t.setStyleSheet("font-size:10px;color:#666;border:none;")
        lbl_v = QLabel("—")
        lbl_v.setStyleSheet("font-size:16px;font-weight:bold;border:none;")
        lbl_v.setAlignment(Qt.AlignCenter)
        lbl_total = QLabel("")
        lbl_total.setStyleSheet("font-size:10px;color:#888;border:none;")
        lbl_total.setAlignment(Qt.AlignCenter)
        lay.addWidget(lbl_t)
        lay.addWidget(lbl_v)
        lay.addWidget(lbl_total)
        frame._lbl_media = lbl_v
        frame._lbl_total = lbl_total
        return frame

    # ── seleção via checkbox ───────────────────────────────
    def _marcar_todos(self, lst):
        lst.blockSignals(True)
        for i in range(lst.count()):
            lst.item(i).setCheckState(Qt.Checked)
        lst.blockSignals(False)
        self._preencher()

    def _desmarcar_todos(self, lst):
        lst.blockSignals(True)
        for i in range(lst.count()):
            lst.item(i).setCheckState(Qt.Unchecked)
        lst.blockSignals(False)
        self._preencher()

    def _popular_listas(self):
        for painel in self._paineis:
            col, lst = painel["col"], painel["lst"]
            antigos_marcados = {
                lst.item(i).text() for i in range(lst.count())
                if lst.item(i).checkState() == Qt.Checked
            }
            lst.blockSignals(True)
            lst.clear()
            if hasattr(self, "_df_full"):
                valores = sorted(v for v in self._df_full[col].dropna().unique() if str(v).strip())
                for v in valores:
                    item = QListWidgetItem(str(v))
                    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                    marcado = (not antigos_marcados) or (str(v) in antigos_marcados)
                    item.setCheckState(Qt.Checked if marcado else Qt.Unchecked)
                    lst.addItem(item)
            lst.blockSignals(False)

    # ── persistência de configuração (config.json) ───────
    def _salvar_config(self):
        cfg_save({"tendencias_config": {
            "f_ano":          self._f_ano.currentText(),
            "f_mes":          self._f_mes.currentText(),
            "periodo_ligado": self._chk_periodo.isChecked(),
            "granularidade":  self._cb_gran.currentText(),
        }})

    def _restaurar_config(self):
        cfg = cfg_load().get("tendencias_config")
        if not cfg:
            return
        self._cfg_pendente = cfg
        if cfg.get("granularidade") in ("Mês", "Ano"):
            self._cb_gran.blockSignals(True)
            self._cb_gran.setCurrentText(cfg["granularidade"])
            self._cb_gran.blockSignals(False)
        if "periodo_ligado" in cfg:
            self._chk_periodo.blockSignals(True)
            self._chk_periodo.setChecked(bool(cfg["periodo_ligado"]))
            self._dt_de.setEnabled(self._chk_periodo.isChecked())
            self._dt_ate.setEnabled(self._chk_periodo.isChecked())
            self._chk_periodo.blockSignals(False)

    def _on_toggle_periodo(self, ligado):
        self._dt_de.setEnabled(ligado)
        self._dt_ate.setEnabled(ligado)
        if ligado:
            hoje = QDate.currentDate()
            self._dt_de.blockSignals(True);  self._dt_de.setDate(hoje);  self._dt_de.blockSignals(False)
            self._dt_ate.blockSignals(True); self._dt_ate.setDate(hoje); self._dt_ate.blockSignals(False)
        self._preencher()

    def atualizar(self):
        if not PANDAS_OK:
            return
        cols, rows = buscar_todos()
        if not rows:
            return
        df = pd.DataFrame(rows, columns=cols)
        df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce").fillna(0)
        df["Mes"]   = pd.to_numeric(df["Mes"],   errors="coerce")
        df["Ano"]   = pd.to_numeric(df["Ano"],   errors="coerce")
        df["_DataDT"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
        df = df[df["Ano"] != 1900]
        self._df_full = df

        anos  = ["(todos)"] + sorted(df["Ano"].dropna().unique().astype(int).astype(str).tolist())
        meses = ["(todos)"] + [f"{i} – {NOMES_MESES[i]}" for i in range(1, 13)]
        cfg_pendente = getattr(self, "_cfg_pendente", None) if not self._cfg_aplicado else None
        for cb, vals, cfg_key in ((self._f_ano, anos, "f_ano"), (self._f_mes, meses, "f_mes")):
            cur = cfg_pendente.get(cfg_key) if cfg_pendente else cb.currentText()
            cb.blockSignals(True)
            cb.clear(); cb.addItems(vals)
            idx = cb.findText(cur) if cur else -1
            cb.setCurrentIndex(idx if idx >= 0 else 0)
            cb.blockSignals(False)
        self._cfg_aplicado = True
        self._popular_listas()
        self._preencher()

    def _contar_periodos(self, df, gran):
        d = df.dropna(subset=["Ano", "Mes"])
        if d.empty:
            return 0
        if gran == "Mês":
            return d[["Ano", "Mes"]].drop_duplicates().shape[0]
        return d["Ano"].drop_duplicates().shape[0]

    def _serie_periodo(self, df, gran):
        if df.empty:
            return []
        d = df.dropna(subset=["Ano", "Mes"]).copy()
        if d.empty:
            return []
        if gran == "Mês":
            d["_key"] = d["Ano"].astype(int) * 100 + d["Mes"].astype(int)
        else:
            d["_key"] = d["Ano"].astype(int)
        return d.groupby("_key")["Valor"].sum().sort_index().tolist()

    def _slope(self, valores):
        n = len(valores)
        if n < 2:
            return None
        xs = range(n)
        mx = sum(xs) / n
        my = sum(valores) / n
        num = sum((x - mx) * (y - my) for x, y in zip(xs, valores))
        den = sum((x - mx) ** 2 for x in xs)
        return num / den if den else 0.0

    def _atualizar_card_destaque(self, df, gran):
        serie = self._serie_periodo(df, gran)
        slope = self._slope(serie)
        if slope is None:
            self._card_tend_saldo._lbl.setText("Dados insuficientes para calcular a tendência")
            cor = "#1565C0"
        else:
            if slope > 0.01:
                sinal = "alta"
            elif slope < -0.01:
                sinal = "queda"
            else:
                sinal = "estável"
            cor = "#1b5e20" if slope >= 0 else "#c62828"
            self._card_tend_saldo._lbl.setText(f"{fmt_valor(slope)} por período  ({sinal})")
        self._card_tend_saldo.setStyleSheet(
            f"QFrame{{background:#f5f5f5;border:2px solid {cor};border-radius:8px;}}")
        self._card_tend_saldo._lbl.setStyleSheet(
            f"font-size:22px;font-weight:bold;color:{cor};border:none;")

    def _atualizar_card_painel(self, painel, df, n_periodos):
        col, lst, card = painel["col"], painel["lst"], painel["card"]
        selecionados = [
            lst.item(i).text() for i in range(lst.count())
            if lst.item(i).checkState() == Qt.Checked
        ]
        df_sel = df[df[col].isin(selecionados)] if selecionados else df.iloc[0:0]
        total = df_sel["Valor"].sum()
        media = total / n_periodos if n_periodos else 0.0
        cor = "#1b5e20" if media >= 0 else "#c62828"
        card._lbl_media.setStyleSheet(f"font-size:16px;font-weight:bold;color:{cor};border:none;")
        card._lbl_media.setText(fmt_valor(media))
        card._lbl_total.setText(f"Total no período: {fmt_valor(total)}")

    def _preencher(self):
        if not hasattr(self, "_df_full"):
            return
        df = self._df_full.copy()
        if self._f_ano.currentText() != "(todos)":
            df = df[df["Ano"] == int(self._f_ano.currentText())]
        mes_sel = self._f_mes.currentText()
        if mes_sel != "(todos)":
            df = df[df["Mes"] == int(mes_sel.split(" – ")[0])]
        if self._chk_periodo.isChecked():
            de  = self._dt_de.date().toPyDate()
            ate = self._dt_ate.date().toPyDate()
            df = df[(df["_DataDT"].dt.date >= de) & (df["_DataDT"].dt.date <= ate)]

        gran = self._cb_gran.currentText()
        self._atualizar_card_destaque(df, gran)
        n_periodos = self._contar_periodos(df, gran)
        for painel in self._paineis:
            self._atualizar_card_painel(painel, df, n_periodos)


# ═══════════════════════════════════════════════════════════
#  ABA CONFIGURAÇÕES
# ═══════════════════════════════════════════════════════════

class AbaConfiguracoes(QWidget):

    def __init__(self, on_db_trocado=None):
        super().__init__()
        self._on_db_trocado = on_db_trocado
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(18)

        # ── Banco de dados ─────────────────────────────────
        grp_db = QGroupBox("Banco de Dados")
        lay_db = QHBoxLayout(grp_db)
        self._ed_db_path = QLineEdit(DB_PATH)
        self._ed_db_path.setReadOnly(True)
        lay_db.addWidget(self._ed_db_path, 1)
        lay_db.addWidget(_btn("Procurar...", "#1565C0", self._procurar_banco, 110))
        root.addWidget(grp_db)

        # ── Senha de acesso ─────────────────────────────────
        grp_senha = QGroupBox("Senha de Acesso ao Programa")
        lay_senha = QFormLayout(grp_senha)
        self._ed_senha1 = QLineEdit(); self._ed_senha1.setEchoMode(QLineEdit.Password)
        self._ed_senha2 = QLineEdit(); self._ed_senha2.setEchoMode(QLineEdit.Password)
        lay_senha.addRow("Nova senha:", self._ed_senha1)
        lay_senha.addRow("Confirmar senha:", self._ed_senha2)
        botoes_senha = QHBoxLayout()
        botoes_senha.addWidget(_btn("Definir senha", "#1b5e20", self._definir_senha, 130))
        botoes_senha.addWidget(_btn("Remover senha", "#c62828", self._remover_senha, 130))
        botoes_senha.addStretch()
        lay_senha.addRow("", botoes_senha)
        self._lbl_status_senha = QLabel()
        lay_senha.addRow("", self._lbl_status_senha)
        root.addWidget(grp_senha)
        self._atualizar_status_senha()

        # ── Backup ───────────────────────────────────────────
        grp_backup = QGroupBox("Backup")
        lay_backup = QHBoxLayout(grp_backup)
        lay_backup.addWidget(QLabel("Salva uma cópia completa do banco de dados atual."))
        lay_backup.addStretch()
        lay_backup.addWidget(_btn("Fazer backup...", "#6A1B9A", self._fazer_backup, 140))
        root.addWidget(grp_backup)

        root.addStretch()

    # ── banco de dados ────────────────────────────────────
    def _procurar_banco(self):
        global DB_PATH
        caminho, _ = QFileDialog.getOpenFileName(
            self, "Selecionar banco de dados", os.path.dirname(DB_PATH),
            "Banco de dados (*.db);;Todos os arquivos (*)")
        if not caminho:
            return
        resp = QMessageBox.question(
            self, "Trocar banco de dados",
            f"Trocar para o banco:\n{caminho}\n\n"
            "O programa vai recarregar todos os dados. Continuar?",
            QMessageBox.Yes | QMessageBox.No)
        if resp != QMessageBox.Yes:
            return
        DB_PATH = caminho
        cfg_save({"db_path": caminho})
        init_db()
        self._ed_db_path.setText(DB_PATH)
        if self._on_db_trocado:
            self._on_db_trocado()
        QMessageBox.information(self, "Banco de dados", "Banco de dados trocado com sucesso.")

    # ── senha de acesso ────────────────────────────────────
    def _atualizar_status_senha(self):
        if cfg_load().get("auth"):
            self._lbl_status_senha.setText("Senha de acesso está ATIVA.")
            self._lbl_status_senha.setStyleSheet("color:#1b5e20; font-weight:bold;")
        else:
            self._lbl_status_senha.setText("Nenhuma senha configurada (acesso livre).")
            self._lbl_status_senha.setStyleSheet("color:#888;")

    def _definir_senha(self):
        s1 = self._ed_senha1.text()
        s2 = self._ed_senha2.text()
        if not s1:
            QMessageBox.warning(self, "Senha", "Digite a nova senha.")
            return
        if s1 != s2:
            QMessageBox.warning(self, "Senha", "As senhas não coincidem.")
            return
        if len(s1) < 4:
            QMessageBox.warning(self, "Senha", "A senha deve ter ao menos 4 caracteres.")
            return
        salt = os.urandom(16).hex()
        cfg_save({"auth": {"salt": salt, "hash": _hash_senha(s1, salt)}})
        self._ed_senha1.clear(); self._ed_senha2.clear()
        self._atualizar_status_senha()
        QMessageBox.information(
            self, "Senha",
            "Senha definida com sucesso. Será exigida na próxima abertura do programa.")

    def _remover_senha(self):
        cfg = cfg_load()
        if not cfg.get("auth"):
            QMessageBox.information(self, "Senha", "Não há senha configurada.")
            return
        resp = QMessageBox.question(self, "Remover senha", "Remover a senha de acesso?",
                                      QMessageBox.Yes | QMessageBox.No)
        if resp != QMessageBox.Yes:
            return
        cfg.pop("auth", None)
        import json
        with open(CFG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
        self._atualizar_status_senha()
        QMessageBox.information(self, "Senha", "Senha removida.")

    # ── backup ─────────────────────────────────────────────
    def _fazer_backup(self):
        pasta = cfg_load().get("backup_dir") or os.path.dirname(DB_PATH)
        sugestao = f"backup_dados_{datetime.datetime.now():%Y%m%d_%H%M%S}.db"
        destino, _ = QFileDialog.getSaveFileName(
            self, "Salvar backup", os.path.join(pasta, sugestao),
            "Banco de dados (*.db)")
        if not destino:
            return
        try:
            shutil.copy2(DB_PATH, destino)
            cfg_save({"backup_dir": os.path.dirname(destino)})
            QMessageBox.information(self, "Backup", f"Backup salvo em:\n{destino}")
        except Exception as e:
            QMessageBox.critical(self, "Backup", f"Erro ao salvar backup:\n{e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Tabela Dinâmica")
        self.setMinimumSize(800, 560)

        # restaurar geometria salva
        cfg = cfg_load().get("janela", {})
        if cfg.get("maximizado"):
            self.resize(cfg.get("largura", 1100), cfg.get("altura", 720))
            self.showMaximized()
        else:
            self.resize(cfg.get("largura", 1100), cfg.get("altura", 720))
            if "x" in cfg and "y" in cfg:
                self.move(cfg["x"], cfg["y"])

        tabs = QTabWidget()
        tabs.setDocumentMode(True)
        tabs.setFont(QFont("Segoe UI", 10))

        self._aba_form       = AbaForm()
        self._aba_import     = AbaImport(self._aba_form)
        self._aba_categorias = AbaCategorias(self._aba_form)
        self._aba_pivot      = AbaPivot()
        self._aba_dash       = AbaDashboard()
        self._aba_tendencias = AbaTendencias()
        self._aba_config     = AbaConfiguracoes(on_db_trocado=self._refresh_tudo)

        tabs.addTab(self._aba_form,       "  Dados  ")
        tabs.addTab(self._aba_import,     "  Importar  ")
        tabs.addTab(self._aba_categorias, "  Categorias  ")
        tabs.addTab(self._aba_pivot,      "  Tabela Dinâmica  ")
        tabs.addTab(self._aba_dash,       "  Dashboard  ")
        tabs.addTab(self._aba_tendencias, "  Tendências  ")
        tabs.addTab(self._aba_config,     "  Configurações  ")
        tabs.currentChanged.connect(self._on_tab)

        self.setCentralWidget(tabs)
        self.statusBar().showMessage("Pronto")

    def closeEvent(self, event):
        maximizado = self.isMaximized()
        if maximizado:
            self.showNormal()  # para capturar tamanho restaurado
        cfg_save({"janela": {
            "maximizado": maximizado,
            "largura":    self.width(),
            "altura":     self.height(),
            "x":          self.x(),
            "y":          self.y(),
        }})
        event.accept()

    def _on_tab(self, idx):
        if idx == 2:
            self._aba_categorias.recarregar()
        if idx == 3:
            self._aba_pivot._gerar()
        if idx == 4:
            self._aba_dash.atualizar()
        if idx == 5:
            self._aba_tendencias.atualizar()

    def _refresh_tudo(self):
        """Recarrega todas as abas após troca de banco de dados."""
        self._aba_form._carregar()
        self._aba_categorias.recarregar()
        self._aba_pivot._gerar()
        self._aba_dash.atualizar()
        self._aba_tendencias.atualizar()


# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":
    _cfg_inicial = cfg_load()
    _db_salvo = _cfg_inicial.get("db_path")
    if _db_salvo and os.path.isfile(_db_salvo):
        DB_PATH = _db_salvo
    init_db()

    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    if not _verificar_login():
        sys.exit(0)

    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
